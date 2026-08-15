# -*- coding: utf-8 -*-
"""第二轮API冒烟：修正payload后的深度业务流验证"""
import json, time, io
import httpx

BASE = "http://127.0.0.1:8000/api/v1"
TS = int(time.time())
results = []

def record(name, passed, detail=""):
    results.append({"test": name, "passed": bool(passed), "detail": str(detail)[:400]})
    print(("PASS" if passed else "FAIL"), "|", name, "|", str(detail)[:180])

def envelope_ok(resp):
    try:
        j = resp.json()
        return isinstance(j, dict) and "code" in j and "success" in j
    except Exception:
        return False

c = httpx.Client(base_url=BASE, timeout=60)
r = c.post("/auth/login", json={"username": "admin", "password": "Admin@12345"})
tok = r.json().get("data", {}).get("access_token")
H = {"Authorization": "Bearer " + tok}
print("login:", r.status_code, "token:", bool(tok))

# 1. 帮扶村创建（snake_case village_name）
vid = None
try:
    vname = f"冒烟村2<scr>&🎉{TS}"
    r = c.post("/supported-villages", json={"village_name": vname, "code": f"SMK{TS}",
               "province": "贵州省", "city": "毕节市", "county": "织金县", "population": 999,
               "support_unit": "测试部队"}, headers=H)
    j = r.json()
    vid = j.get("data", {}).get("id")
    record("R2-01 创建帮扶村", r.status_code in (200, 201) and vid, f"id={vid} status={r.status_code}")
except Exception as e:
    record("R2-01 创建帮扶村", False, e)

# 2. SQL注入（正确参数名 village_name）
try:
    r = c.get("/supported-villages", params={"village_name": "' OR '1'='1", "code": "' OR '1'='1", "page": 1, "page_size": 10}, headers=H)
    j = r.json()
    n = len(j.get("data", {}).get("items", []))
    record("R2-02 SQL注入参数化", r.status_code == 200 and n == 0, f"matched={n}")
except Exception as e:
    record("R2-02 SQL注入参数化", False, e)

# 3. 年度数据 + 变更历史
if vid:
    try:
        r = c.post(f"/supported-villages/{vid}/yearly/2026/income", json={"household_income": 8888.8, "notes": "R2年度数据"}, headers=H)
        r2 = c.get(f"/supported-villages/{vid}/yearly/2026", headers=H)
        income = r2.json().get("data", {}).get("income")
        r3 = c.get(f"/supported-villages/{vid}/change-history", headers=H)
        hist = r3.json().get("data", [])
        record("R2-03 年度数据+变更历史", r.status_code == 200 and income is not None and len(hist) > 0,
               f"income={income is not None} history={len(hist)}")
    except Exception as e:
        record("R2-03 年度数据+变更历史", False, e)

# 4. 软删除 + include_deleted
if vid:
    try:
        r = c.delete(f"/supported-villages/{vid}", headers=H)
        j = r.json()
        r2 = c.get("/supported-villages", params={"page": 1, "page_size": 100}, headers=H)
        items = r2.json().get("data", {}).get("items", [])
        gone = all(i.get("id") != vid for i in items)
        r3 = c.get("/supported-villages", params={"page": 1, "page_size": 100, "include_deleted": True}, headers=H)
        items3 = r3.json().get("data", {}).get("items", [])
        back = any(i.get("id") == vid and (i.get("is_deleted") or not i.get("isActive")) for i in items3)
        record("R2-04 软删除+include_deleted", r.status_code == 200 and gone and back,
               f"default_hidden={gone} include_deleted_visible={back} approval_task={j.get('data', {}).get('approval_task_id')}")
    except Exception as e:
        record("R2-04 软删除+include_deleted", False, e)

# 5. 审批流：提交(change_data) → 驳回空原因 → 驳回带原因 → 通知
fid = None
try:
    r = c.post("/funds", json={"name": f"审批流经费{TS}", "planned_amount": 30000}, headers=H)
    fid = r.json().get("data", {}).get("id")
    r = c.post("/approval/submit", json={"entity_type": "fund", "entity_id": fid, "title": f"经费审批{fid}", "change_data": {"name": f"审批流经费{TS}"}}, headers=H)
    j = r.json()
    task_id = j.get("data", {}).get("id") or j.get("data", {}).get("task_id")
    record("R2-05 提交审批", r.status_code in (200, 201) and task_id, f"task={task_id} status={r.status_code}")
    if task_id:
        r = c.post(f"/approval/tasks/{task_id}/reject", json={"reason": ""}, headers=H)
        record("R2-06 空驳回原因拦截", r.status_code in (400, 422), f"status={r.status_code}")
        r = c.post(f"/approval/tasks/{task_id}/reject", json={"reason": "预算超标，请重新编制"}, headers=H)
        record("R2-07 驳回(带原因)", r.status_code == 200, f"status={r.status_code} body={r.text[:120]}")
        r = c.get("/messages", params={"page": 1, "page_size": 30}, headers=H)
        txt = json.dumps(r.json().get("data", {}).get("items", []), ensure_ascii=False)
        record("R2-08 驳回消息通知", "驳回" in txt or "预算" in txt, f"found_in_messages={('驳回' in txt) or ('预算' in txt)}")
except Exception as e:
    record("R2-05..08 审批流", False, e)

# 6. 审批通过 + 状态流转
fid2 = None
try:
    r = c.post("/funds", json={"name": f"审批通过经费{TS}", "planned_amount": 20000}, headers=H)
    fid2 = r.json().get("data", {}).get("id")
    r = c.post("/approval/submit", json={"entity_type": "fund", "entity_id": fid2, "title": f"经费审批{fid2}", "change_data": {}}, headers=H)
    task_id = r.json().get("data", {}).get("id") or r.json().get("data", {}).get("task_id")
    r = c.post(f"/approval/tasks/{task_id}/approve", json={"opinion": "同意拨付"}, headers=H)
    r2 = c.get(f"/funds/{fid2}", headers=H)
    status = r2.json().get("data", {}).get("status")
    record("R2-09 审批通过+状态更新", r.status_code == 200 and status in ("approved", "已审批", "approval_passed", "allocated"),
           f"status={status} approve_resp={r.text[:120]}")
except Exception as e:
    record("R2-09 审批通过+状态更新", False, e)

# 7. 预算预警（amount 字段）
fid3 = None
try:
    r = c.post("/funds", json={"name": f"预算预警经费{TS}", "planned_amount": 10000}, headers=H)
    fid3 = r.json().get("data", {}).get("id")
    r = c.put(f"/funds/{fid3}", json={"amount": 10000, "approved_amount": 10000, "used_amount": 9200}, headers=H)
    time.sleep(0.5)
    r1 = c.post("/reminders/scan", headers=H)
    time.sleep(0.5)
    def count_bw():
        r = c.get("/reminders", params={"page": 1, "page_size": 200}, headers=H)
        items = r.json().get("data", {}).get("items", [])
        return sum(1 for i in items if i.get("type") == "budget_warning" and i.get("entity_id") == fid3)
    c1 = count_bw()
    r2 = c.post("/reminders/scan", headers=H)
    time.sleep(0.5)
    c2 = count_bw()
    record("R2-10 预算预警92%+幂等", r1.status_code == 200 and c1 >= 1 and c2 == c1, f"first={c1} second={c2}")
except Exception as e:
    record("R2-10 预算预警92%+幂等", False, e)

# 8. 里程碑列表格式
try:
    r = c.post("/projects", json={"name": f"里程碑项目{TS}", "start_date": "2026-02-01", "status": "pending",
                                  "budget": 100, "category": "产业", "location": "贵州", "priority": "low"}, headers=H)
    pid = r.json().get("data", {}).get("id")
    c.post(f"/projects/{pid}/milestones", json={"name": "里程碑A", "planned_date": "2026-07-01"}, headers=H)
    r = c.get(f"/projects/{pid}/milestones", headers=H)
    j = r.json()
    data = j.get("data")
    is_list = isinstance(data, list)
    record("R2-11 里程碑列表格式", r.status_code == 200 and is_list and len(data) >= 1, f"data_is_list={is_list} n={len(data) if is_list else data}")
except Exception as e:
    record("R2-11 里程碑列表格式", False, e)

# 9. 413 大文件（png 类型）
try:
    big = b"\x89PNG\r\n\x1a\n" + b"\x00" * (11 * 1024 * 1024)
    r = c.post("/files/upload", files={"file": ("big.png", big, "image/png")}, headers=H)
    record("R2-12 11MB文件413", r.status_code == 413, f"status={r.status_code} body={r.text[:120]}")
except Exception as e:
    record("R2-12 11MB文件413", False, e)

# 10. 转账凭证创建
try:
    r = c.post("/fund-lifecycle/transfer-vouchers", json={"voucher_no": f"V{TS}", "direction": "military_to_local",
               "amount": 5000, "transfer_date": "2026-08-15", "project_id": 1}, headers=H)
    record("R2-13 转账凭证创建", r.status_code in (200, 201), f"status={r.status_code} body={r.text[:150]}")
    r = c.get("/fund-lifecycle/transfer-vouchers", params={"page": 1, "page_size": 10}, headers=H)
    record("R2-14 转账凭证列表", envelope_ok(r), r.text[:120])
except Exception as e:
    record("R2-13/14 转账凭证", False, e)

# 11. 工作日志
try:
    r = c.post("/work-logs", json={"content": f"冒烟工作日志{TS}", "log_date": "2026-08-15"}, headers=H)
    record("R2-15 工作日志创建", r.status_code in (200, 201), f"status={r.status_code} body={r.text[:120]}")
    r = c.get("/work-logs", params={"page": 1, "page_size": 5}, headers=H)
    record("R2-16 工作日志列表信封", envelope_ok(r), r.text[:120])
except Exception as e:
    record("R2-15/16 工作日志", False, e)

# 12. Excel 导入帮扶村（模板+1坏行）
try:
    r = c.get("/supported-villages/import-template", headers=H)
    tpl = r.content
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(tpl))
    ws = wb.active
    hdr = [cell.value for cell in ws[1]]
    record("R2-17 导入模板下载", r.status_code == 200 and len(tpl) > 100, f"headers={hdr}")
    # 在模板第2行填一行合法数据、第3行填一行超长/错误数据
    name_col = hdr.index("village_name") if "village_name" in hdr else (hdr.index("村名") if "村名" in hdr else 0)
    code_col = hdr.index("code") if "code" in hdr else 1
    ws.cell(row=2, column=name_col + 1, value=f"导入测试村{TS}")
    ws.cell(row=2, column=code_col + 1, value=f"IMP{TS}")
    ws.cell(row=3, column=name_col + 1, value="x" * 300)  # 超长村名
    ws.cell(row=3, column=code_col + 1, value=f"IMPB{TS}")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = c.post("/supported-villages/import", files={"file": (f"import{TS}.xlsx", buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, headers=H)
    record("R2-18 Excel导入(1坏行)", r.status_code == 200, f"status={r.status_code} body={r.text[:400]}")
except Exception as e:
    record("R2-17/18 Excel导入", False, e)

# 13. 受助学生导入
try:
    r = c.get("/schools", params={"page": 1, "page_size": 1}, headers=H)
    sid = r.json().get("data", {}).get("items", [{}])[0].get("id")
    if sid:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in [["student_name", "grade", "amount"], ["张三", "三年级", 500], ["李四", "三年级", 800]]:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = c.post(f"/schools/{sid}/scholarship-students/import", files={"file": ("stu.xlsx", buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, headers=H)
        record("R2-19 受助学生导入", r.status_code in (200, 201), f"status={r.status_code} body={r.text[:250]}")
        r = c.get(f"/schools/{sid}/scholarship-students", headers=H)
        record("R2-20 受助学生列表", r.status_code == 200, r.text[:150])
    else:
        record("R2-19/20 受助学生", False, "no school")
except Exception as e:
    record("R2-19/20 受助学生", False, e)

# 14. 消息未读数
try:
    r = c.get("/messages/unread-count", headers=H)
    record("R2-21 未读数接口", r.status_code == 200, r.text[:120])
except Exception as e:
    record("R2-21 未读数接口", False, e)

summary = {"total": len(results), "passed": sum(1 for x in results if x["passed"]),
           "failed": sum(1 for x in results if not x["passed"]), "results": results}
print("===SUMMARY===")
print(json.dumps(summary, ensure_ascii=False))
