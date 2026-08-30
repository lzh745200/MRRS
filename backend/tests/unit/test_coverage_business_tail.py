"""
定向覆盖测试 — 补齐 5 个文件的缺失分支（业务尾段）。

- app/api/v1/supported_village.py   (23 lines)
- app/api/v1/policy.py              (22 lines)
- app/services/approval_workflow_service.py (22 lines)
- app/api/v1/funds.py               (21 lines)
- app/api/v1/permission_package.py  (20 lines)

原则：不改源码与已有测试；优先直接调用内部函数以精确定位分支。
"""
import asyncio
import io
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from fastapi import HTTPException

import openpyxl


# ============================================================================
# 通用辅助
# ============================================================================


def _chained_db():
    """返回链式 MagicMock 会话：query/filter/order_by/offset/limit 均返回自身。"""
    db = MagicMock()
    db.query.return_value = db
    db.filter.return_value = db
    db.order_by.return_value = db
    db.offset.return_value = db
    db.limit.return_value = db
    db.options.return_value = db
    db.join.return_value = db
    return db


def _admin_user():
    return SimpleNamespace(
        id=1, username="admin", role="admin", is_superuser=True, is_active=True,
        organization_id=1, permissions_list=["*"], full_name="管理员",
    )


def _non_admin_user():
    return SimpleNamespace(
        id=2, username="bob", role="user", is_superuser=False, is_active=True,
        organization_id=2, permissions_list=[], full_name="普通用户",
    )


def _aws_mock_db():
    """审批服务链式 mock 会话（query 返回独立 query mock）。"""
    db = MagicMock()
    q = MagicMock()
    db.query.return_value = q
    q.options.return_value = q
    q.filter.return_value = q
    q.order_by.return_value = q
    q.offset.return_value = q
    q.limit.return_value = q
    q.all.return_value = []
    q.count.return_value = 0
    q.first.return_value = None
    return db


# ============================================================================
# app/api/v1/supported_village.py
# ============================================================================


class TestSupportedVillageTail:
    def test_apply_village_approval_result_noop(self):
        from app.api.v1.supported_village import _apply_village_approval_result
        assert _apply_village_approval_result(None, None) is None

    def test_find_village_header_row_fallback(self):
        from app.api.v1.supported_village import _find_village_header_row, _FIELD_NAMES
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["foo", "bar"])  # 无 village_name 标签 → 回退分支
        idx, col_map = _find_village_header_row(ws)
        assert idx == 1
        assert col_map == {name: i for i, name in enumerate(_FIELD_NAMES)}

    def test_process_import_row_bool_parsing(self):
        from app.api.v1.supported_village import _process_import_row
        db = _chained_db()
        db.first.return_value = None
        col_map = {"village_name": 0, "is_three_regions": 1}
        ok, err = _process_import_row(("测试村", "是"), col_map, db, 2, None)
        assert ok is True and err is None
        assert db.add.call_count == 1
        village = db.add.call_args[0][0]
        assert village.is_three_regions is True

    def test_list_villages_boolean_filters(self):
        from app.api.v1.supported_village import list_villages
        db = _chained_db()
        db.count.return_value = 0
        db.all.return_value = []
        with patch("app.api.v1.supported_village.apply_scope_filter",
                   side_effect=lambda q, *a, **k: q):
            result = asyncio.run(list_villages(
                page=1, page_size=20,
                is_three_regions=True, is_ethnic_area=True, is_key_county=True,
                include_deleted=False, current_user=_admin_user(), db=db,
            ))
        assert result["code"] == 200
        assert result["data"]["total"] == 0

    # ── _coerce_section_value ──

    def test_coerce_none_value(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import VillagePopulation
        assert _coerce_section_value(VillagePopulation, "total_households", None) is None
        assert _coerce_section_value(VillagePopulation, "total_households", "   ") is None

    def test_coerce_unknown_attr(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import VillagePopulation
        assert _coerce_section_value(VillagePopulation, "no_such_attr", "x") == "x"

    def test_coerce_boolean(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import SupportedVillage
        assert _coerce_section_value(SupportedVillage, "is_three_regions", "是") is True
        assert _coerce_section_value(SupportedVillage, "is_three_regions", "0") is False

    def test_coerce_integer(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import VillagePopulation
        assert _coerce_section_value(VillagePopulation, "total_households", "12.7") == 12

    def test_coerce_float(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import VillageIncome
        assert _coerce_section_value(VillageIncome, "per_capita_income", "3.5") == 3.5

    def test_coerce_invalid_returns_none(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import VillagePopulation
        assert _coerce_section_value(VillagePopulation, "total_households", "abc") is None

    def test_coerce_string_strip(self):
        from app.api.v1.supported_village import _coerce_section_value
        from app.models.supported_village import SupportedVillage
        assert _coerce_section_value(SupportedVillage, "village_name", " 测试 ") == "测试"

    # ── _import_section_sheet ──

    def test_import_section_empty_rows(self):
        from app.api.v1.supported_village import _import_section_sheet
        from app.models.supported_village import VillagePopulation
        wb = openpyxl.Workbook()
        result = _import_section_sheet(wb.active, VillagePopulation, 1, 2025, _chained_db())
        assert result == {"imported": 0, "failed": 0}

    def test_import_section_branches(self):
        from app.api.v1.supported_village import _import_section_sheet
        from app.models.supported_village import VillagePopulation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "总户数", "总人数"])   # 表头（含 None 单元格）
        ws.append([None, None, None])          # 空行
        ws.append(["示例行", 1, 2])            # 示例行
        ws.append(["x", None, None])           # 无有效数据行
        ws.append(["y", 100, 200])             # 有效行
        with patch("app.api.v1.supported_village._save_section_data"):
            result = _import_section_sheet(ws, VillagePopulation, 1, 2025, _chained_db())
        assert result == {"imported": 1, "failed": 1}

    def test_import_section_exception(self):
        from app.api.v1.supported_village import _import_section_sheet
        from app.models.supported_village import VillagePopulation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["总户数", "总人数"])
        ws.append([100, 200])
        with patch("app.api.v1.supported_village._save_section_data",
                   side_effect=RuntimeError("boom")):
            result = _import_section_sheet(ws, VillagePopulation, 1, 2025, _chained_db())
        assert result == {"imported": 0, "failed": 1}

    def test_import_all_sections_unrecognized_sheet(self):
        from app.api.v1.supported_village import import_all_sections_data
        db = _chained_db()
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"  # 未识别工作表 → 跳过
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        class _FakeFile:
            async def read(self):
                return buf.getvalue()

        with patch("app.api.v1.supported_village._get_village_or_404", return_value=MagicMock()):
            with patch("app.api.v1.supported_village.safe_commit"):
                result = asyncio.run(import_all_sections_data(
                    1, None, _FakeFile(), _admin_user(), db,
                ))
        assert result["code"] == 200
        assert result["data"]["sheets"] == 0


# ============================================================================
# app/api/v1/policy.py
# ============================================================================


class TestPolicyTail:
    def test_apply_policy_approval_result_noop(self):
        from app.api.v1.policy import _apply_policy_approval_result
        assert _apply_policy_approval_result(None, None) is None

    def test_build_pdf_font_fallback_and_detail_branches(self):
        from app.api.v1.policy import _build_policies_pdf
        from reportlab.pdfbase import pdfmetrics

        p = SimpleNamespace(
            level="national", status="active", category="local",
            title="政策A", code="P001", issuing_authority="机关",
            issue_date=None, effective_date=None, content=None, keywords="kw",
            file_path="att.pdf",
        )
        real_register = pdfmetrics.registerFont

        def _fake_register(font):
            if getattr(font, "fontName", None) == "STSong-Light":
                raise Exception("no cjk font")
            return real_register(font)

        with patch.object(pdfmetrics, "registerFont", side_effect=_fake_register):
            data = _build_policies_pdf([p])
        assert isinstance(data, bytes) and len(data) > 0

    # ── batch-delete 校验分支 ──

    def test_batch_delete_ids_not_list(self):
        from app.api.v1.policy import batch_delete_policies
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(batch_delete_policies({"ids": "not-a-list"}, _admin_user(), _chained_db()))
        assert exc_info.value.status_code == 422

    def test_batch_delete_ids_non_int(self):
        from app.api.v1.policy import batch_delete_policies
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(batch_delete_policies({"ids": [1, "two"]}, _admin_user(), _chained_db()))
        assert exc_info.value.status_code == 422

    def test_batch_delete_ids_bool_or_nonpositive(self):
        from app.api.v1.policy import batch_delete_policies
        with pytest.raises(HTTPException):
            asyncio.run(batch_delete_policies({"ids": [True]}, _admin_user(), _chained_db()))
        with pytest.raises(HTTPException):
            asyncio.run(batch_delete_policies({"ids": [0]}, _admin_user(), _chained_db()))

    def test_batch_delete_ids_too_many(self):
        from app.api.v1.policy import batch_delete_policies
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(batch_delete_policies(
                {"ids": list(range(1, 1002))}, _admin_user(), _chained_db()))
        assert exc_info.value.status_code == 422

    # ── 列表非管理员过滤 ──

    def test_get_policies_non_admin_filter(self):
        from app.api.v1.policy import get_policies
        db = _chained_db()
        db.count.return_value = 0
        db.all.return_value = []
        result = asyncio.run(get_policies(
            skip=None, limit=None, category=None, organization_level=None,
            search=None, order_by=None, order_desc=None, year=None,
            document_code=None, page=None, page_size=None, keyword=None,
            level=None, status=None,
            current_user=_non_admin_user(), db=db,
        ))
        assert result["code"] == 200
        assert result["data"]["total"] == 0

    def test_get_related_policies_non_admin_filter(self):
        from app.api.v1.policy import get_related_policies
        db = _chained_db()
        policy = SimpleNamespace(id=5, category="local", created_by=999)
        db.first.return_value = policy
        db.all.return_value = []
        result = asyncio.run(get_related_policies(
            5, current_user=_non_admin_user(), db=db))
        assert result["code"] == 200

    def test_get_policy_forbidden(self):
        from app.api.v1.policy import get_policy
        db = _chained_db()
        policy = SimpleNamespace(id=5, status="draft", created_by=999)
        db.first.return_value = policy
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(get_policy(5, current_user=_non_admin_user(), db=db))
        assert exc_info.value.status_code == 403

    # ── FTS 失败分支（create/update）──

    def test_create_policy_fts_failure(self, real_db_session):
        from app.api.v1.policy import PolicyCreateRequest, create_policy
        db = real_db_session
        data = PolicyCreateRequest(title="FTS失败测试", content="正文", category="local")
        with patch("app.services.policy_fts_service.ensure_fts_table",
                   side_effect=Exception("ensure fail")),              patch("app.services.policy_fts_service.sync_policy_to_fts",
                   side_effect=Exception("sync fail")):
            result = asyncio.run(create_policy(data, _admin_user(), db))
        assert result["code"] == 200
        assert result["data"]["title"] == "FTS失败测试"

    def test_update_policy_fts_failure(self, real_db_session):
        from app.api.v1.policy import PolicyUpdateRequest, update_policy
        from app.models.policy import Policy
        db = real_db_session
        p = Policy(title="旧标题", status="draft", category="local")
        db.add(p)
        db.commit()
        db.refresh(p)
        pid = p.id
        data = PolicyUpdateRequest(title="新标题")
        with patch("app.services.policy_fts_service.ensure_fts_table",
                   side_effect=Exception("ensure fail")),              patch("app.services.policy_fts_service.sync_policy_to_fts",
                   side_effect=Exception("sync fail")):
            result = asyncio.run(update_policy(pid, data, _admin_user(), db))
        assert result["code"] == 200
        assert result["data"]["title"] == "新标题"


# ============================================================================
# app/services/approval_workflow_service.py
# ============================================================================


class TestApprovalWorkflowServiceTail:
    def test_submit_approval_role_node(self):
        from app.services.approval_workflow_service import ApprovalWorkflowService
        db = _aws_mock_db()
        q = db.query.return_value
        first_node = MagicMock()
        first_node.approver_type = "role"
        first_node.approver_id = "admin"
        wf = MagicMock()
        wf.id = 10
        wf.nodes = [first_node]
        q.options.return_value.filter.return_value.first.return_value = wf

        svc = ApprovalWorkflowService(db=db)
        with patch.object(svc, "_resolve_role_approver_id", return_value=7):
            result = svc.submit_approval("fund", 1, 100, title="t")
        assert result is not None
        assert result.current_approver_id == 7

    def test_get_tasks_with_count_filters(self):
        from app.services.approval_workflow_service import ApprovalWorkflowService
        db = _aws_mock_db()
        svc = ApprovalWorkflowService(db=db)
        result = svc.get_tasks_with_count(
            status="pending", completed=False,
            date_from="2026-08-01", date_to="2026-08-02",
        )
        assert result == {"items": [], "total": 0}

    def test_get_tasks_with_count_invalid_dates(self):
        from app.services.approval_workflow_service import ApprovalWorkflowService
        db = _aws_mock_db()
        svc = ApprovalWorkflowService(db=db)
        result = svc.get_tasks_with_count(date_from="not-a-date", date_to="also-bad")
        assert result == {"items": [], "total": 0}

    def test_auto_approve_all_pending_exception(self):
        from app.services.approval_workflow_service import ApprovalWorkflowService
        db = _aws_mock_db()
        q = db.query.return_value
        task = MagicMock()
        task.id = 1
        q.all.return_value = [task]
        svc = ApprovalWorkflowService(db=db)
        with patch.object(svc, "approve_task", side_effect=RuntimeError("boom")):
            result = svc.auto_approve_all_pending(10)
        assert result == {"total_pending": 1, "approved": 0, "failed": 1}
        db.rollback.assert_called()

    def test_get_task_diff_non_dict(self):
        from app.services.approval_workflow_service import ApprovalWorkflowService
        db = _aws_mock_db()
        svc = ApprovalWorkflowService(db=db)
        task = MagicMock()
        task.id = 5
        task.entity_type = "fund"
        task.entity_id = 10
        task.change_data = "not-dict"
        task.original_data = ["also-not-dict"]
        with patch.object(svc, "get_task", return_value=task):
            result = svc.get_task_diff(5)
        assert result["changed"] == {}
        assert result["original"] == {}

    def test_submit_entity_change_approval_priority(self):
        from app.services.approval_workflow_service import (
            ApprovalWorkflowService, submit_entity_change_approval,
        )
        db = MagicMock()
        task = MagicMock()
        task.id = 99
        with patch.object(ApprovalWorkflowService, "submit_approval", return_value=task):
            with patch("app.services.approval_workflow_service.safe_commit") as mock_sc:
                tid = submit_entity_change_approval(
                    db, entity_type="fund", entity_id=1, submitter_id=10,
                    title="t", priority=5,
                )
        assert tid == 99
        assert task.priority == 5
        mock_sc.assert_called_once_with(db)


# ============================================================================
# app/api/v1/funds.py
# ============================================================================


class TestFundsTail:
    def test_resolve_fund_approval_tasks_reject(self):
        from app.api.v1.funds import _resolve_fund_approval_tasks
        from app.models.approval import ApprovalStatus
        db = _chained_db()
        task_in = MagicMock()
        task_in.id = 1
        db.all.return_value = [task_in]
        rejected = MagicMock()
        rejected.status = ApprovalStatus.REJECTED.value
        service = MagicMock()
        service.reject_task.return_value = rejected
        fund = SimpleNamespace(id=5)
        operator = SimpleNamespace(id=10)
        with patch("app.api.v1.funds.ApprovalWorkflowService", return_value=service):
            count = _resolve_fund_approval_tasks(db, fund, "reject", operator)
        assert count == 1
        service.reject_task.assert_called_once_with(
            1, 10, "经费板块直接驳回", standalone=True)

    def test_apply_fund_approval_result_skips_non_terminal(self):
        from app.api.v1.funds import _apply_fund_approval_result
        task = SimpleNamespace(status="pending")
        assert _apply_fund_approval_result(MagicMock(), task) is None

    def test_apply_fund_approval_result_no_fund(self):
        from app.api.v1.funds import _apply_fund_approval_result
        db = _chained_db()
        db.first.return_value = None
        task = SimpleNamespace(status="approved", entity_id=5)
        assert _apply_fund_approval_result(db, task) is None

    def test_apply_fund_approval_result_current_approver(self):
        from app.api.v1.funds import _apply_fund_approval_result
        db = _chained_db()
        fund = SimpleNamespace(id=5, status="pending")
        db.first.return_value = fund
        approver = SimpleNamespace(full_name="张三", username="zs")
        task = SimpleNamespace(status="approved", entity_id=5, current_approver=approver,
                               current_approver_id=10, id=1)
        _apply_fund_approval_result(db, task)
        assert fund.status == "approved"
        assert fund.approved_by == "张三"
        assert db.add.call_count == 1

    def test_apply_fund_approval_result_resolve_user(self):
        from app.api.v1.funds import _apply_fund_approval_result
        db = _chained_db()
        fund = SimpleNamespace(id=5, status="pending")
        approver = SimpleNamespace(full_name="李四", username="ls")
        db.first.side_effect = [fund, approver]
        task = SimpleNamespace(status="approved", entity_id=5, current_approver=None,
                               current_approver_id=10, id=1)
        _apply_fund_approval_result(db, task)
        assert fund.status == "approved"
        assert fund.approved_by == "李四"

    def test_update_fund_no_changes(self):
        from app.api.v1.funds import FundUpdate, update_fund
        db = _chained_db()
        fund = SimpleNamespace(id=1, status="pending")
        with patch("app.api.v1.funds._get_fund_or_404", return_value=fund):
            result = update_fund(1, FundUpdate(), _admin_user(), db)
        assert result["code"] == 200
        assert result["message"] == "更新成功"
        assert "data" not in result

    def test_fund_statistics_year_filter(self):
        from app.api.v1.funds import fund_statistics_overview
        db = _chained_db()
        row = SimpleNamespace(
            total_count=5, total_amount=1000.0, pending_count=1, approved_count=1,
            allocated_count=1, in_use_count=1, completed_count=1,
            used_amount=100.0, allocated_amount=200.0,
        )
        exec_result = MagicMock()
        exec_result.one.return_value = row
        db.execute.return_value = exec_result
        db.first.return_value = (100, 50)
        with patch("app.api.v1.funds.apply_scope_filter",
                   side_effect=lambda stmt, *a, **k: stmt):
            result = fund_statistics_overview(year=2025, current_user=_admin_user(), db=db)
        assert result["code"] == 200
        assert result["data"]["year"] == 2025

    def test_fund_approval_flow(self):
        from app.api.v1.funds import fund_approval_flow
        db = _chained_db()
        db.all.return_value = []
        fund = SimpleNamespace(id=5, status="approved", approved_by="张三",
                               approval_date=None)
        with patch("app.api.v1.funds._get_fund_or_404", return_value=fund):
            result = fund_approval_flow(5, _admin_user(), db)
        assert result["code"] == 200
        assert result["data"]["current_status"] == "approved"
        assert len(result["data"]["nodes"]) == 6


# ============================================================================
# app/api/v1/permission_package.py
# ============================================================================


class TestPermissionPackageTail:
    def test_optional_current_user_no_auth(self):
        from app.api.v1.permission_package import _optional_current_user
        assert _optional_current_user(None) is None

    def test_optional_current_user_empty_payload(self):
        from app.api.v1.permission_package import _optional_current_user
        with patch("app.core.security.decode_token", return_value={}):
            assert _optional_current_user("Bearer sometoken") is None

    def test_optional_current_user_valid_token(self):
        from app.api.v1.permission_package import _optional_current_user
        user = SimpleNamespace(username="admin")
        session = MagicMock()
        q = MagicMock()
        session.query.return_value = q
        q.filter.return_value = q
        q.first.return_value = user
        with patch("app.core.security.decode_token", return_value={"sub": "admin"}):
            with patch("app.core.database.SessionLocal", return_value=session):
                result = _optional_current_user("Bearer sometoken")
        assert result is user
        session.close.assert_called_once()

    def test_optional_current_user_http_exception(self):
        from app.api.v1.permission_package import _optional_current_user
        with patch("app.core.security.decode_token",
                   side_effect=HTTPException(status_code=401, detail="bad")):
            assert _optional_current_user("Bearer sometoken") is None

    def test_download_traversal_bad_name(self):
        from app.api.v1.permission_package import download_permission_package
        with patch("app.utils.paths.get_uploads_path", return_value="/uploads/pp"):
            with pytest.raises(HTTPException) as exc_info:
                download_permission_package("../evil.zip", _admin_user())
        assert exc_info.value.status_code == 400
        assert exc_info.value.detail == "非法文件名"

    def test_download_traversal_realpath(self):
        import app.api.v1.permission_package as pp
        from app.api.v1.permission_package import download_permission_package

        def _fake_realpath(p, strict=False):
            if str(p).endswith("foo.zip"):
                return "/outside/foo.zip"
            return "/uploads/pp"

        # 源码已迁移 get_uploads_path → get_runtime_uploads_path（W6 路径双源收口）
        with patch("app.utils.paths.get_runtime_uploads_path", return_value="/uploads/pp"):
            with patch.object(pp.os.path, "realpath", side_effect=_fake_realpath):
                with pytest.raises(HTTPException) as exc_info:
                    download_permission_package("foo.zip", _admin_user())
        assert exc_info.value.status_code == 400
        assert exc_info.value.detail == "非法文件路径"
