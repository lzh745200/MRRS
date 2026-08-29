"""Tests for batch_import_optimizer.py — 100% coverage target."""

import pytest
from unittest.mock import Mock, patch
from io import BytesIO


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_excel_bytes():
    """Create a small valid Excel file in memory using openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "city"])
    ws.append(["Alice", "30", "Beijing"])
    ws.append(["Bob", "25", "Shanghai"])
    ws.append(["Charlie", "35", "Shenzhen"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def mock_db():
    return Mock()


# ---------------------------------------------------------------------------
# read_excel_fast
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# _read_excel_fallback
# ---------------------------------------------------------------------------

class TestReadExcelFallback:
    """Cover all branches of _read_excel_fallback."""

    def test_normal_read(self):
        """Normal sheet with headers and data rows."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["col_a", "col_b"])
        ws.append(["x", "1"])
        ws.append(["y", "2"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        assert headers == ["col_a", "col_b"]
        assert len(rows) == 2
        assert rows[0] == {"col_a": "x", "col_b": "1"}

    def test_empty_sheet_stop_iteration(self):
        """Empty sheet → StopIteration → return [], []."""
        from openpyxl import Workbook
        wb = Workbook()
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        assert headers == []
        assert rows == []

    def test_empty_rows_skipped(self):
        """Rows where all values are None should be skipped."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append([None, None])   # should be skipped
        ws.append(["v1", "v2"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        assert len(rows) == 1
        assert rows[0] == {"a": "v1", "b": "v2"}

    def test_row_fewer_values_than_headers(self):
        """Row with fewer cells than headers — missing cells are None → empty string."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["h1", "h2", "h3"])
        ws.append(["only"])   # only 1 value for 3 headers
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        assert headers == ["h1", "h2", "h3"]
        assert len(rows) == 1
        # openpyxl pads to max column count → missing become None → ""
        assert rows[0] == {"h1": "only", "h2": "", "h3": ""}

    def test_row_more_values_than_headers(self):
        """Row with more cells than headers — headers are padded with col_N;
        the i < len(headers) guard prevents index errors."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["h1", "h2"])
        ws.append(["a", "b", "c", "d"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        # openpyxl pads header row to max column count → 2 new col_N headers
        assert len(headers) == 4
        assert headers[0] == "h1"
        assert headers[1] == "h2"
        assert headers[2] == "col_2"
        assert headers[3] == "col_3"
        assert len(rows) == 1
        assert rows[0] == {"h1": "a", "h2": "b", "col_2": "c", "col_3": "d"}

    def test_header_none_generates_col_name(self):
        """When a header cell is None/empty, a col_N placeholder is generated."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["good", None, "also_good"])
        ws.append(["1", "2", "3"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        rows, headers = _read_excel_fallback(buf.getvalue())
        assert len(headers) == 3
        # the second header should be "col_1" (index 1)
        assert headers[0] == "good"
        assert headers[1] == "col_1"
        assert headers[2] == "also_good"
        assert rows[0]["col_1"] == "2"


# ---------------------------------------------------------------------------
# validate_rows
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# batch_insert_optimized
# ---------------------------------------------------------------------------

