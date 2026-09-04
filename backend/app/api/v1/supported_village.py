"""帮扶村管理 API 路由"""

# 数据权限过滤已迁移到 app.core.data_scope_adapter.apply_scope_filter()
# 支持组织树展开（org_children 含下级组织），与 school.py 行为一致

import io
import json
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import AliasChoices, BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.core.response import ok_list, success_response
from app.core.security import get_current_user
from app.models.user import User
from app.models.audit import AuditAction
from app.utils.common import dict_keys_to_camel, StringHelper
from app.models.supported_village import (
    SupportedVillage,
    VillagePopulation,
    VillageIncome,
    ForceInvestment,
    IndustrySupport,
    InfrastructureImprovement,
    PartyBuildingSupport,
    MedicalSupport,
    ConsumptionSupport,
    EmploymentSupport,
    EducationSupport,
    VillageCommitteeInfo,
    VillageCommitteeMember,
)
from app.core.data_permission import check_record_access
from app.core.data_scope_adapter import apply_scope_filter
from app.api.v1.deps import enforce_admin_include_deleted, build_viewable_because
from app.schemas.supported_village import SupportedVillageCreate, SupportedVillageUpdate
from app.core.transaction import safe_commit, savepoint
from app.services.approval_workflow_service import (
    ApprovalWorkflowService,
    submit_entity_change_approval,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/supported-villages", tags=["帮扶村管理"])


def _apply_village_approval_result(db: Session, task) -> None:
    """审批终态回写帮扶村（注册到 ApprovalWorkflowService）

    帮扶村无 pending 状态门（创建即生效），审批通过后无字段需回写；
    此处保留处理器接口，便于后续扩展（如 transition_status 审批门）。
    """
    _ = (db, task)


ApprovalWorkflowService.register_entity_apply_handler("supported_village", _apply_village_approval_result)


class BatchDeleteRequest(BaseModel):
    ids: List[int]
    confirm_password: Optional[str] = ""


class PurgeRequest(BaseModel):
    """彻底删除（回收站清空单条）— 二次密码确认防误删"""
    confirm_password: str = ""


class YearCopyRequest(BaseModel):
    """年度数据复制请求。

    H2d（功能恒 422 修复）：字段名为 camelCase(fromYear/toYear)且必填，但
    CamelToSnakeMiddleware 会把请求体键名转为 from_year/to_year，Pydantic 找
    不到必填 camelCase 字段 → POST /yearly/copy 恒抛 422，年度复制 100% 不可用。
    照搬 TransitionFundingItem(H2b) 样板：populate_by_name + AliasChoices 同时
    兼容中间件转换后的 snake_case 与未经中间件的 camelCase。
    """

    model_config = {"populate_by_name": True}

    fromYear: int = Field(validation_alias=AliasChoices("fromYear", "from_year"))
    toYear: int = Field(validation_alias=AliasChoices("toYear", "to_year"))


class YearlySectionData(BaseModel):
    """年度区块数据 — 字段由各 section 动态决定，路由层过滤安全属性"""
    year: Optional[int] = None
    model_config = {"extra": "allow"}


# ── 年度数据 Section → Model 映射 ──
_SECTION_MODEL: Dict[str, Any] = {
    "population": VillagePopulation,
    "income": VillageIncome,
    "force-investment": ForceInvestment,
    "industry": IndustrySupport,
    "infrastructure": InfrastructureImprovement,
    "party-building": PartyBuildingSupport,
    "medical": MedicalSupport,
    "consumption": ConsumptionSupport,
    "employment": EmploymentSupport,
    "education": EducationSupport,
    "committee": VillageCommitteeInfo,
}

# ── 导入导出列定义 ──
_IMPORT_COLUMNS = [
    ("village_name", "帮扶村名称"),
    ("department", "部门单位"),
    ("support_unit", "帮扶单位"),
    ("province", "省"),
    ("city", "市"),
    ("county", "县/市"),
    ("township", "乡镇"),
    ("village_category", "村庄类别"),
    ("is_three_regions", "是否三区三州"),
    ("is_key_county", "是否重点帮扶县"),
    ("is_revitalization_tier", "是否振兴梯队"),
    ("longitude", "经度"),
    ("latitude", "纬度"),
    ("altitude", "海拔"),
    ("area_sq_km", "面积(km²)"),
    ("households", "户数"),
    ("description", "备注"),
]

# 预计算，避免重复 destructure
_FIELD_NAMES = [col[0] for col in _IMPORT_COLUMNS]
_HEADER_NAMES = [col[1] for col in _IMPORT_COLUMNS]

# 表头标签 → 模型字段（覆盖官方模板 VILLAGE_FIELDS 19 列 + 旧平铺格式别名）
# 官方模板：表头第 6 行、首列"序号"、表头带必填 * 前缀
_IMPORT_LABEL_MAP = {
    "各部门各单位": "department",
    "部门单位": "department",
    "具体帮扶单位": "support_unit",
    "帮扶单位": "support_unit",
    "定点帮扶村": "village_name",
    "帮扶村名称": "village_name",
    "省": "province",
    "市": "city",
    "县/市": "county",
    "县 / 市": "county",
    "乡镇": "township",
    "地区范围": "region_scope",
    "是否属于三区三州": "is_three_regions",
    "是否三区三州": "is_three_regions",
    "是否属于边疆地区": "is_border_area",
    "是否属于民族地区": "is_ethnic_area",
    "是否属于革命地区": "is_revolutionary_area",
    "是否属于160个国家乡村振兴重点帮扶县": "is_key_county",
    "是否重点帮扶县": "is_key_county",
    "是否振兴梯队": "is_revitalization_tier",
    "省级乡村振兴示范创建对象": "is_provincial_demo",
    "百村示范创建对象": "is_hundred_village_demo",
    "梯次振兴发展对象": "is_tiered_development",
    "是否跨省": "is_cross_province",
    "是否跨市": "is_cross_city",
    "是否跨大单位协作帮扶": "is_cross_unit_cooperation",
    "是否纳入总盘子": "is_in_overall_plan",
    "2021年以来获得的国家或省级表彰": "honors",
    "经度": "longitude",
    "纬度": "latitude",
}

_BOOL_IMPORT_FIELDS = frozenset({
    "is_three_regions", "is_border_area", "is_ethnic_area", "is_revolutionary_area",
    "is_key_county", "is_revitalization_tier", "is_provincial_demo",
    "is_hundred_village_demo", "is_tiered_development", "is_cross_province",
    "is_cross_city", "is_cross_unit_cooperation", "is_in_overall_plan",
})


class _ImportRowRejected(Exception):
    """M6：导入行业务性拒绝（重名/校验失败）——用于触发 SAVEPOINT 回滚该行。"""


# 年度数据辅助函数中需要跳过的元数据列
_SKIP_COLUMNS = frozenset({"id", "supported_village_id", "year", "created_at", "updated_at"})


# ═══════════════════════════════════════════════════════════════
#  辅助函数
# ═══════════════════════════════════════════════════════════════

def _get_section_data(db: Session, model: Any, village_id: int, year: int) -> Optional[Dict]:
    row = db.query(model).filter(
        model.supported_village_id == village_id, model.year == year
    ).first()
    if not row:
        return None
    result = {}
    for col in model.__table__.columns:
        if col.name in _SKIP_COLUMNS:
            continue
        val = getattr(row, col.name)
        result[col.name] = val
    # 村委会板块：加载成员列表
    if model is VillageCommitteeInfo:
        members = db.query(VillageCommitteeMember).filter(
            VillageCommitteeMember.committee_info_id == row.id
        ).all()
        result["members"] = [
            dict_keys_to_camel({
                "name": m.name,
                "position": m.position,
                "phone": m.phone,
                "is_veteran": m.is_veteran,
                "remark": m.remark,
            })
            for m in members
        ]
    return dict_keys_to_camel(result)


def _copy_section_data(db: Session, model: Any, village_id: int, from_year: int, to_year: int) -> bool:
    src = db.query(model).filter(
        model.supported_village_id == village_id, model.year == from_year
    ).first()
    if not src:
        return False
    existing = db.query(model).filter(
        model.supported_village_id == village_id, model.year == to_year
    ).first()
    if existing:
        return False
    new_row = model()
    new_row.supported_village_id = village_id
    new_row.year = to_year
    for col in model.__table__.columns:
        if col.name in _SKIP_COLUMNS:
            continue
        setattr(new_row, col.name, getattr(src, col.name, None))
    db.add(new_row)
    return True


def _save_section_data(db: Session, model: Any, village_id: int, year: int, data: dict):
    row = db.query(model).filter(
        model.supported_village_id == village_id, model.year == year
    ).first()
    if not row:
        row = model()
        row.supported_village_id = village_id
        row.year = year
        db.add(row)
    # 处理村委会成员子表
    members_data = data.pop("members", None)
    # 将 camelCase 键转为 snake_case 以匹配模型属性名
    snake_data = {}
    for key, value in data.items():
        if key in _SKIP_COLUMNS:
            continue
        snake_key = StringHelper.to_snake_case(key) if hasattr(key, 'upper') else key
        snake_data[snake_key] = value
    for key, value in snake_data.items():
        if hasattr(row, key):
            setattr(row, key, value)
        else:
            # H6：未知字段不再静默跳过——显式记录 warning，避免 schema/模型字段漂移
            # 导致"以为保存成功实际未落库"（如旧 VillageIncome.total_income 幽灵字段）
            logger.warning(
                "_save_section_data: 字段 '%s' 在 %s 模型上不存在，已跳过"
                "（village_id=%s, year=%s）",
                key, getattr(model, "__name__", model), village_id, year,
            )
    if members_data is not None and model is VillageCommitteeInfo:
        # 新建行需先 flush 拿到主键，成员才能正确外键关联
        if row.id is None:
            db.flush()
        # 清除旧成员，写入新成员
        db.query(VillageCommitteeMember).filter(
            VillageCommitteeMember.committee_info_id == row.id
        ).delete()
        for m in members_data:
            if isinstance(m, dict):
                # H2c：CamelToSnakeMiddleware 会递归转换数组内键名
                # （isVeteran→is_veteran），原代码硬编码读 camelCase 的 isVeteran
                # 恒取不到 → 退役军人标记静默落库 False。与顶层字段一致，对成员
                # 子表键先做 to_snake_case 归一化，再读 snake_case（保留 camelCase 兜底
                # 以兼容未经中间件的直接调用）。
                nm = {}
                for k, v in m.items():
                    nk = StringHelper.to_snake_case(k) if hasattr(k, "upper") else k
                    nm[nk] = v
                member = VillageCommitteeMember(
                    committee_info_id=row.id,
                    supported_village_id=village_id,
                    name=nm.get("name", ""),
                    position=nm.get("position", ""),
                    phone=nm.get("phone", ""),
                    is_veteran=bool(nm.get("is_veteran", nm.get("isVeteran", False))),
                    remark=nm.get("remark", ""),
                )
                db.add(member)
    return row


def _village_to_diff_dict(village: SupportedVillage) -> Dict[str, Any]:
    """帮扶村字段级 Diff 快照（排除审计元数据列）"""
    skip = frozenset({"id", "organization_id", "created_by", "created_at", "updated_at", "is_active"})
    result = {}
    for col in SupportedVillage.__table__.columns:
        if col.name in skip:
            continue
        result[col.name] = getattr(village, col.name, None)
    return dict_keys_to_camel(result)


def _record_village_change(
    db: Session,
    action: Any,
    current_user: User,
    village: SupportedVillage,
    old_data: Optional[Dict[str, Any]] = None,
    new_data: Optional[Dict[str, Any]] = None,
    detail: Optional[str] = None,
):
    """记录帮扶村字段级变更历史（与项目/经费一致的审计留痕）"""
    from app.services.audit_enhancement_service import AuditEnhancementService

    AuditEnhancementService.log_entity_changes(
        db,
        action,
        current_user,
        "supported_village",
        str(village.id),
        old_data,
        new_data,
        detail=detail,
    )


def _get_village_or_404(db: Session, village_id: int, current_user: User = None) -> SupportedVillage:
    """根据 ID 获取帮扶村，不存在时抛 404；存在但跨组织时抛 403（数据隔离）。"""
    from sqlalchemy.orm import selectinload

    village = (
        db.query(SupportedVillage)
        .filter(SupportedVillage.id == village_id)
        .options(selectinload(SupportedVillage.organization))
        .first()
    )
    if not village:
        raise HTTPException(status_code=404, detail="帮扶村不存在")
    # 数据权限校验：非本组织/非本人创建且非管理员 → 403（区分"不存在"与"越权"）
    if current_user is not None and not check_record_access(
        village, current_user, owner_field="created_by", dept_field="organization_id"
    ):
        raise HTTPException(status_code=403, detail="无权访问该帮扶村")
    return village


async def _invalidate_village_cache():
    """清除帮扶村列表缓存，确保写操作后立即可见"""
    import os
    if os.environ.get("PYTEST_CURRENT_TEST"):
        return
    try:
        from app.core.cache import get_cache_service
        _cache = await get_cache_service()
        await _cache.delete_by_prefix("villages:list:")
    except Exception as e:
        logger.debug("清理帮扶村列表缓存失败: %s", e)


def _find_village_header_row(ws) -> tuple:
    """探测表头行，返回 (header_row_idx, col_map: {field_name: col_index})。

    官方模板：第 1-5 行为装饰标题区，第 6 行表头（带 * 前缀），第 7 行示例行。
    用户自制平铺文件：第 1 行表头。按表头标签驱动列映射，避免位置错位。
    探测失败回退：假定第 1 行表头、按 _IMPORT_COLUMNS 位置映射（旧行为）。
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        col_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).lstrip("*").strip()
            field = _IMPORT_LABEL_MAP.get(label)
            if field and field not in col_map:
                col_map[field] = col_idx
        if "village_name" in col_map:
            return idx, col_map
    # 回退：旧平铺格式（无表头识别时按位置映射）
    return 1, {name: i for i, name in enumerate(_FIELD_NAMES)}


def _process_import_row(
    row: tuple, col_map: Dict[str, int], db: Session, row_idx: int, current_user=None
):
    """处理单行导入数据（表头驱动列映射）。返回 (success, error_msg)"""
    values: Dict[str, Any] = {}
    for field_name, col_idx in col_map.items():
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None and isinstance(val, str):
            val = val.strip()
        if field_name in _BOOL_IMPORT_FIELDS:
            # 空单元格保持 None（不写 False），有值才解析布尔
            if val is None or (isinstance(val, str) and not val):
                val = None
            else:
                val = str(val).strip() in ("是", "1", "True", "true", "yes", "Y")
        values[field_name] = val
    if not values.get("village_name"):
        return False, f"第{row_idx}行: 帮扶村名称不能为空"
    # 字段长度校验（与手工创建 schema 对齐，防止超长数据绕过校验入库）
    _import_max_lengths = {
        "village_name": 200, "sequence_no": 50, "department": 100,
        "support_unit": 200, "province": 50, "city": 50, "county": 50,
        "township": 50, "support_contact": 50, "support_contact_phone": 20,
    }
    for _f, _max in _import_max_lengths.items():
        _v = values.get(_f)
        if isinstance(_v, str) and len(_v) > _max:
            return False, f"第{row_idx}行: 字段 '{_f}' 长度超过 {_max} 字符限制（实际 {len(_v)} 字符）"
    existing = db.query(SupportedVillage).filter(
        SupportedVillage.village_name == values["village_name"],
        SupportedVillage.county == values.get("county"),
        SupportedVillage.is_active == True,  # noqa: E712
    ).first()
    if existing:
        return False, f"第{row_idx}行: 帮扶村 '{values['village_name']}' 已存在，跳过"
    village = SupportedVillage(**{k: v for k, v in values.items() if v is not None})
    # 归属字段：与手工创建一致，保证导入者在数据权限范围内可见/可编辑
    if current_user is not None:
        village.organization_id = getattr(current_user, "organization_id", None)
        village.created_by = getattr(current_user, "id", None)
    db.add(village)
    return True, None


# ═══════════════════════════════════════════════════════════════
#  列表 & 筛选选项（无路径参数，必须在 /{village_id} 之前注册）
# ═══════════════════════════════════════════════════════════════


@router.get("")
async def list_villages(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: Optional[str] = None,
    department: Optional[str] = None,
    county: Optional[str] = None,
    is_revitalization_tier: Optional[bool] = None,
    is_three_regions: Optional[bool] = None,
    is_ethnic_area: Optional[bool] = None,
    is_key_county: Optional[bool] = None,
    year_start: Optional[int] = None,
    with_summary: bool = Query(False),
    isRevitalizationTier: Optional[bool] = None,  # 兼容旧参数名
    include_deleted: bool = Depends(enforce_admin_include_deleted),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村列表（分页、筛选）"""
    # 缓存：帮扶村日级更新，TTL 120s
    # 跳过测试环境——模块级缓存单例会在测试间泄漏状态
    import hashlib
    import json
    import os
    # 兼容旧 camelCase 参数名（新代码统一 snake_case）——缓存 key 与过滤共用
    tier = is_revitalization_tier if is_revitalization_tier is not None else isRevitalizationTier
    _ckey = None
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        from app.core.cache import get_cache_service
        _cache = await get_cache_service()
        try:
            _key_data = json.dumps(
                [keyword, department, county, tier, is_three_regions, is_ethnic_area,
                 is_key_county, include_deleted, year_start, with_summary],
                default=str,
            ).encode()
            _org_id = getattr(current_user, "organization_id", None) or 0
            _hash = hashlib.md5(_key_data, usedforsecurity=False).hexdigest()
            _ckey = f"villages:list:{_org_id}:{page}:{page_size}:{_hash}"
            _cached = await _cache.get(_ckey)
            if _cached is not None:
                return _cached
        except (TypeError, ValueError):
            _ckey = None

    # include_deleted 已由 enforce_admin_include_deleted 依赖收敛：非管理员自动降级为 False
    from sqlalchemy.orm import selectinload

    # organization 关系为 lazy="noload", 列表序列化需读取组织名, 必须显式预加载
    query = db.query(SupportedVillage).options(selectinload(SupportedVillage.organization))
    # 数据权限过滤（统一使用 data_scope_adapter，支持 org_children 含下级组织）
    query = apply_scope_filter(query, current_user, SupportedVillage, db=db)

    # 默认过滤软删记录（is_active=False），include_deleted=True 时显示全部
    if not include_deleted:
        query = query.filter(SupportedVillage.is_active == True)  # noqa: E712

    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            SupportedVillage.village_name.ilike(like)
            | SupportedVillage.support_unit.ilike(like)
        )
    if department:
        query = query.filter(SupportedVillage.department == department)
    if county:
        query = query.filter(SupportedVillage.county == county)
    if year_start is not None:
        query = query.filter(func.strftime("%Y", SupportedVillage.created_at) == str(year_start))
    # 兼容旧 camelCase 参数名（新代码统一 snake_case）
    if tier is not None:
        query = query.filter(SupportedVillage.is_revitalization_tier == bool(tier))
    if is_three_regions is not None:
        query = query.filter(SupportedVillage.is_three_regions == bool(is_three_regions))
    if is_ethnic_area is not None:
        query = query.filter(SupportedVillage.is_ethnic_area == bool(is_ethnic_area))
    if is_key_county is not None:
        query = query.filter(SupportedVillage.is_key_county == bool(is_key_county))

    summary = None
    if with_summary:
        # H1：KPI「总投入」口径改为 SupportedVillage.transition_fund_military_total +
        # transition_fund_local_total 求和——这是录入表单 save_transition_funding 的唯一
        # 写入目标；此前误读从未被任何生产路径写入的 SupportFunding 子表，导致
        # summary["total_investment"] 恒为 0。复用已应用 apply_scope_filter(数据隔离) +
        # is_active==True 过滤的 query，一并消除原 fund_agg 全局无隔离求和的越权隐患。
        fund_agg = (
            query.with_entities(
                func.coalesce(func.sum(SupportedVillage.transition_fund_military_total), 0),
                func.coalesce(func.sum(SupportedVillage.transition_fund_local_total), 0),
            ).first()
        )
        agg = (
            query.with_entities(
                func.count(SupportedVillage.id),
                func.count(func.distinct(SupportedVillage.county)),
                func.count(func.distinct(SupportedVillage.department)),
            ).first()
        )
        summary = {
            "total": int(agg[0] or 0),
            "total_investment": round(float(fund_agg[0] or 0) + float(fund_agg[1] or 0), 4),
            "county_count": int(agg[1] or 0),
            "department_count": int(agg[2] or 0),
        }

    total = query.count()
    # Model-level lazy="selectin" on SupportedVillage relationships prevents
    # N+1 queries when accessing yearly-data backrefs during iteration/serialization.
    items = (
        query
        .order_by(SupportedVillage.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # 返回统一 envelope：{code:200, data:{items,total,page,page_size}, message}
    result = ok_list(
        items=[
            v.to_dict() if hasattr(v, "to_dict") else {"id": v.id, "village_name": v.village_name}
            for v in items
        ],
        total=total,
        page=page,
        page_size=page_size,
        extra={"summary": summary} if summary is not None else None,
    )
    if _ckey is not None:
        await _cache.set(_ckey, result, ttl=120)
    return result


@router.get("/filter-options")
async def get_filter_options(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取筛选选项（部门、县市列表）

    数据隔离：仅返回当前用户数据范围内的部门/县市，避免跨组织泄露。
    """
    # 基础查询应用数据范围过滤，再投影到 department/county 列
    base_query = db.query(SupportedVillage)
    base_query = apply_scope_filter(base_query, current_user, SupportedVillage, db=db)
    base_query = base_query.filter(SupportedVillage.is_active == True)  # noqa: E712

    departments = (
        base_query.with_entities(SupportedVillage.department)
        .filter(SupportedVillage.department.isnot(None))
        .distinct()
        .all()
    )
    counties = (
        base_query.with_entities(SupportedVillage.county)
        .filter(SupportedVillage.county.isnot(None))
        .distinct()
        .all()
    )
    return success_response(
        data={
            "departments": [d[0] for d in departments if d[0]],
            "counties": [c[0] for c in counties if c[0]],
            "years": sorted(
                {
                    int(row[0])
                    for row in base_query.with_entities(
                        func.substr(SupportedVillage.created_at, 1, 4)
                    )
                    .distinct()
                    .all()
                    if row and row[0] and str(row[0]).isdigit()
                },
                reverse=True,
            ),
        }
    )


@router.get("/options/dropdown")
async def get_village_dropdown(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村下拉选项（id + name + county，供前端 Select 使用）

    数据隔离：仅返回当前用户数据范围内的帮扶村，避免跨组织泄露。
    """
    query = db.query(
        SupportedVillage.id,
        SupportedVillage.village_name,
        SupportedVillage.county,
    )
    query = apply_scope_filter(query, current_user, SupportedVillage, db=db)
    villages = (
        query.filter(SupportedVillage.is_active == True)  # noqa: E712
        .order_by(SupportedVillage.id)
        .all()
    )
    items = [
        {"id": v[0], "name": v[1], "county": v[2] or ""}
        for v in villages
    ]
    return ok_list(items, len(items))


@router.get("/import-template")
async def download_import_template():
    """下载导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_village_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''supported_village_import_template.xlsx"},
    )


@router.post("/import")
async def import_villages(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入帮扶村"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的文件")
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请检查文件格式")
    header_row, col_map = _find_village_header_row(ws)
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件中没有数据行，请至少添加一行数据")
    imported = 0
    errors = []
    for row_idx, row in enumerate(rows, start=header_row + 1):
        if not any(row):
            continue
        # 跳过模板示例行（行尾标注"← 示例行（导入时自动跳过）"）与说明/页脚行
        if any(
            isinstance(c, str) and ("示例行" in c or "填写说明" in c or "帮扶管理信息系统 v" in c)
            for c in row
        ):
            continue
        try:
            # M6：每行包裹 SAVEPOINT（app.core.transaction.savepoint）——某行 db.add 后
            # 抛异常时仅回滚该行的脏对象，不随最终 safe_commit 落库产生半条数据。
            with savepoint(db):
                success, error_msg = _process_import_row(row, col_map, db, row_idx, current_user)
                if not success:
                    # 业务性跳过（重名/名称为空/超长）：主动抛出行内异常触发 SAVEPOINT 回滚
                    raise _ImportRowRejected(error_msg or f"第{row_idx}行: 校验失败")
            imported += 1
        except Exception as e:
            # savepoint 已将原始异常包装为 DatabaseError(raise ... from e)；
            # 取 __cause__ 还原可读的行内错误信息，避免消息嵌套重复
            cause = getattr(e, "__cause__", None) or e
            msg = str(cause)
            errors.append(msg if msg.startswith("第") else f"第{row_idx}行: {msg}")
    safe_commit(db)
    await _invalidate_village_cache()
    # 数据变更自动创建审批任务：批量导入进入待审批板块（审计留痕）
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=0,
        submitter_id=current_user.id,
        title=f"帮扶村批量导入：成功 {imported} 条，跳过 {len(errors)} 条",
        change_data={"imported": imported, "failed": len(errors), "errors": errors[:20]},
    )
    return success_response(
        data={"imported": imported, "failed": len(errors), "errors": errors,
              "approval_task_id": approval_task_id},
        message=f"成功导入 {imported} 条记录" + (f"，{len(errors)} 条跳过" if errors else ""),
    )


@router.post("/batch-delete")
async def batch_delete_villages(
    data: BatchDeleteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量软删帮扶村（仅可删除当前用户有权访问的记录，二次密码确认防误删）"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="请提供要删除的ID列表")
    # 二次确认：校验当前用户密码（单机共用电脑防误删，与组织删除对齐）
    from app.core.security import verify_password
    if not data.confirm_password or not verify_password(
        data.confirm_password, getattr(current_user, "hashed_password", "") or ""
    ):
        raise HTTPException(status_code=400, detail="二次确认失败：密码不正确")
    query = db.query(SupportedVillage).filter(SupportedVillage.id.in_(data.ids))
    # 数据权限过滤：仅允许操作本组织/本人创建的记录，防止跨组织批量删除
    query = apply_scope_filter(query, current_user, SupportedVillage, db=db)
    deleted_count = query.update(
        {"is_active": False, "deleted_at": datetime.now(timezone.utc)},
        synchronize_session=False,
    )
    safe_commit(db)
    await _invalidate_village_cache()
    # 数据变更自动创建审批任务：批量删除进入待审批板块（审计留痕）
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=0,
        submitter_id=current_user.id,
        title=f"帮扶村批量删除：{deleted_count} 条",
        change_data={"deleted": True, "deleted_count": deleted_count, "ids": data.ids},
    )
    # 单机防丢失：批量删除后触发一次即时备份
    if deleted_count:
        from app.services.immediate_backup import trigger_immediate_backup

        trigger_immediate_backup(description=f"批量删除帮扶村{deleted_count}条后备份", delay=1.0)
    return success_response(
        data={"deleted": deleted_count, "approval_task_id": approval_task_id},
        message=f"已删除 {deleted_count} 条记录",
    )


# ═══════════════════════════════════════════════════════════════
#  单个帮扶村 CRUD（/{village_id} 必须在所有显式路径之后注册）
# ═══════════════════════════════════════════════════════════════


@router.get("/{village_id}")
async def get_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村详情（含已软删记录，管理员可见时附带 viewableBecause 元数据）"""
    village = _get_village_or_404(db, village_id, current_user)
    data = village.to_dict() if hasattr(village, "to_dict") else {"id": village.id}
    # 管理员查看已软删记录时附带可见性元数据，便于前端审计展示
    data["viewableBecause"] = build_viewable_because(current_user, village)
    return success_response(data=data)


@router.post("")
async def create_village(
    data: SupportedVillageCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建帮扶村"""
    village = SupportedVillage(**data.model_dump())
    # 强制写入数据归属字段，确保新记录纳入组织隔离体系
    village.organization_id = getattr(current_user, "organization_id", None)
    village.created_by = current_user.id
    db.add(village)
    safe_commit(db)
    db.refresh(village)
    await _invalidate_village_cache()
    _record_village_change(
        db, AuditAction.CREATE, current_user, village, new_data=_village_to_diff_dict(village),
        detail=f"创建帮扶村: {village.village_name}",
    )
    # 数据变更自动创建审批任务（Requirement 3.2）：帮扶村新增进入待审批板块
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=village.id,
        submitter_id=current_user.id,
        title=f"帮扶村新增：{village.village_name}",
        change_data=_village_to_diff_dict(village),
    )
    return success_response(data={"id": village.id, "approval_task_id": approval_task_id},
                            message="创建成功")


@router.put("/{village_id}")
async def update_village(
    village_id: int,
    data: SupportedVillageUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新帮扶村（过渡状态变更需管理员权限 + 写入审计日志）"""
    village = _get_village_or_404(db, village_id, current_user)
    update_dict = data.model_dump(exclude_unset=True)
    old_snapshot = _village_to_diff_dict(village)

    # ── 过渡状态变更强制审批检查 ──
    if "transition_status" in update_dict:
        from app.api.v1.deps import require_manager_role
        require_manager_role(current_user)
        old_status = village.transition_status or "none"
        new_status = update_dict["transition_status"]

        # 记录审计日志
        from app.utils.audit_logger import AuditLogger
        AuditLogger.log(
            action="village_transition_change",
            user_id=current_user.id,
            username=current_user.username,
            resource_type="supported_village",
            resource_id=village_id,
            details={
                "old_transition_status": old_status,
                "new_transition_status": new_status,
                "village_name": village.village_name,
            },
        )

    for key, value in update_dict.items():
        if key == "id":
            continue
        if hasattr(village, key):
            setattr(village, key, value)
        else:
            # H6：未知字段不再静默跳过——显式记录 warning，防止字段漂移导致数据未落库
            logger.warning(
                "update_village: 字段 '%s' 在 SupportedVillage 模型上不存在，已跳过"
                "（village_id=%s）",
                key, village_id,
            )
    safe_commit(db)
    await _invalidate_village_cache()
    _record_village_change(
        db, AuditAction.UPDATE, current_user, village,
        old_data=old_snapshot, new_data=_village_to_diff_dict(village),
        detail=f"更新帮扶村: {village.village_name}",
    )
    # 数据变更自动创建审批任务（Requirement 3.2）：帮扶村修改进入待审批板块（含变更对比）
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=village.id,
        submitter_id=current_user.id,
        title=f"帮扶村变更：{village.village_name}",
        change_data=_village_to_diff_dict(village),
        original_data=old_snapshot,
    )
    return success_response(data={"id": village.id, "approval_task_id": approval_task_id},
                            message="更新成功")


@router.delete("/{village_id}")
async def delete_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """软删帮扶村（置 is_active=False，保留关联数据以便恢复/审计）"""
    village = _get_village_or_404(db, village_id, current_user)

    village.is_active = False
    village.deleted_at = datetime.now(timezone.utc)
    safe_commit(db)
    await _invalidate_village_cache()
    # 数据变更自动创建审批任务：帮扶村删除进入待审批板块
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=village.id,
        submitter_id=current_user.id,
        title=f"帮扶村删除：{village.village_name}",
        change_data={"deleted": True, "village_name": village.village_name},
        original_data=_village_to_diff_dict(village),
    )
    return success_response(data={"id": village.id, "approval_task_id": approval_task_id},
                            message="删除成功")


# ── 回收站：恢复 / 彻底删除（仅管理员） ──


def _require_village_in_recycle_bin(village: SupportedVillage) -> None:
    if village.is_active:
        raise HTTPException(status_code=400, detail="该记录不在回收站中")


@router.get("/{village_id}/purge/preview")
async def preview_purge_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """彻底删除预览：返回将级联删除的关联数据统计（仅管理员）"""
    from app.core.permission_utils import require_admin

    require_admin(current_user)
    village = _get_village_or_404(db, village_id, current_user)
    _require_village_in_recycle_bin(village)
    from app.services.village_cascade_delete_service import VillageCascadeDeleteService

    refs = VillageCascadeDeleteService(db).check_village_references(village_id)
    return success_response(
        data={
            "id": village.id,
            "village_name": village.village_name,
            **refs,
        },
    )


@router.post("/{village_id}/restore")
async def restore_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从回收站恢复软删帮扶村（is_active 置回 True，仅管理员）"""
    from app.core.permission_utils import require_admin

    require_admin(current_user)
    village = _get_village_or_404(db, village_id, current_user)
    _require_village_in_recycle_bin(village)
    old_snapshot = _village_to_diff_dict(village)
    village.is_active = True
    village.deleted_at = None
    safe_commit(db)
    await _invalidate_village_cache()
    _record_village_change(
        db, AuditAction.UPDATE, current_user, village,
        old_data=old_snapshot, new_data=_village_to_diff_dict(village),
        detail=f"从回收站恢复帮扶村: {village.village_name}",
    )
    approval_task_id = submit_entity_change_approval(
        db,
        entity_type="supported_village",
        entity_id=village.id,
        submitter_id=current_user.id,
        title=f"帮扶村恢复：{village.village_name}",
        change_data={"restored": True, "village_name": village.village_name},
    )
    return success_response(
        data={"id": village.id, "approval_task_id": approval_task_id},
        message="恢复成功",
    )


@router.post("/{village_id}/purge")
async def purge_village(
    village_id: int,
    data: PurgeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """彻底删除回收站中的帮扶村：物理删除并级联清除全部子表数据（仅管理员 + 密码确认）"""
    from app.core.permission_utils import require_admin

    require_admin(current_user)
    # 二次密码确认（与批量软删一致，防误删）
    from app.core.security import verify_password

    if not data.confirm_password or not verify_password(
        data.confirm_password, getattr(current_user, "hashed_password", "") or ""
    ):
        raise HTTPException(status_code=400, detail="二次确认失败：密码不正确")
    village = _get_village_or_404(db, village_id, current_user)
    _require_village_in_recycle_bin(village)
    village_name = village.village_name
    from app.services.village_cascade_delete_service import VillageCascadeDeleteService

    stats = VillageCascadeDeleteService(db).delete_village_cascade(village_id)
    if not stats.get("success"):
        raise HTTPException(status_code=404, detail=stats.get("message", "彻底删除失败"))
    await _invalidate_village_cache()
    # 审计留痕（村庄本体已物理删除，仅写操作日志）
    from app.utils.audit_logger import AuditLogger

    AuditLogger.log(
        action="village_purge",
        user_id=current_user.id,
        username=current_user.username,
        resource_type="supported_village",
        resource_id=village_id,
        details={
            "village_name": village_name,
            "deleted_records": stats.get("deleted_records"),
            "cascade_details": stats.get("details"),
        },
    )
    # 单机防丢失：彻底删除后触发一次即时备份
    from app.services.immediate_backup import trigger_immediate_backup

    trigger_immediate_backup(
        description=f"彻底删除帮扶村[{village_name}]及{stats.get('deleted_records', 0)}条关联数据后备份",
        delay=1.0,
    )
    return success_response(
        data={
            "id": village_id,
            "deleted_records": stats.get("deleted_records", 0),
            "details": stats.get("details", {}),
        },
        message=f"已彻底删除 [{village_name}] 及 {stats.get('deleted_records', 0)} 条关联数据",
    )


# ═══════════════════════════════════════════════════════════════
#  年度数据
# ═══════════════════════════════════════════════════════════════


@router.get("/{village_id}/yearly/{year}")
async def get_yearly_data(
    village_id: int,
    year: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村某年度全部数据（所有section）"""
    _get_village_or_404(db, village_id, current_user)
    # 统一 camelCase 键名（内层已由 _get_section_data→dict_keys_to_camel 处理）
    # camelCase 为主键名，snake_case 保留向后兼容（计划 2027-01-01 移除）
    result = {"villageId": village_id, "village_id": village_id, "year": year}
    for section, model in _SECTION_MODEL.items():
        data = _get_section_data(db, model, village_id, year)
        result[section] = data if data else None
    return success_response(data=result, message="ok")


@router.post("/{village_id}/yearly/copy")
async def copy_year_data(
    village_id: int,
    data: YearCopyRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """将某年的全部年度数据复制到另一年"""
    _get_village_or_404(db, village_id, current_user)
    if data.fromYear == data.toYear:
        raise HTTPException(status_code=400, detail="来源年份和目标年份不能相同")
    copied = 0
    for model in _SECTION_MODEL.values():
        if _copy_section_data(db, model, village_id, data.fromYear, data.toYear):
            copied += 1
    safe_commit(db)
    await _invalidate_village_cache()
    return success_response(message=f"年度数据复制成功，已复制 {copied} 个数据组")


@router.post("/{village_id}/yearly/{year}/{section}")
async def save_yearly_section(
    village_id: int,
    year: int,
    section: str,
    data: YearlySectionData,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存帮扶村某年度某个section的数据"""
    if section not in _SECTION_MODEL:
        raise HTTPException(status_code=400, detail=f"未知的数据分类: {section}")
    village = _get_village_or_404(db, village_id, current_user)
    model = _SECTION_MODEL[section]
    old_data = _get_section_data(db, model, village_id, year)
    _save_section_data(db, model, village_id, year, data.model_dump())
    safe_commit(db)
    _record_village_change(
        db, AuditAction.UPDATE, current_user, village,
        old_data=old_data, new_data=_get_section_data(db, model, village_id, year),
        detail=f"年度数据保存: {section} {year}年",
    )
    await _invalidate_village_cache()
    return success_response(message=f"保存成功: {section}")


@router.delete("/{village_id}/yearly/{year}/{section}")
async def delete_yearly_section(
    village_id: int,
    year: int,
    section: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除某板块某年度数据（物理删除 + 审计留痕；T028）"""
    if section not in _SECTION_MODEL:
        raise HTTPException(status_code=400, detail=f"未知年度数据板块: {section}")
    village = _get_village_or_404(db, village_id, current_user)
    model = _SECTION_MODEL[section]
    old_data = _get_section_data(db, model, village_id, year)
    if not old_data:
        raise HTTPException(status_code=404, detail=f"{section} {year} 年度数据不存在")
    row = (
        db.query(model)
        .filter(model.supported_village_id == village_id, model.year == year)
        .first()
    )
    if row:
        db.delete(row)
        safe_commit(db)
    _record_village_change(
        db, AuditAction.DELETE, current_user, village,
        old_data=old_data, new_data=None,
        detail=f"年度数据删除: {section} {year}年",
    )
    await _invalidate_village_cache()
    return success_response(message=f"已删除: {section} {year}年")


@router.get("/{village_id}/change-history")
async def get_village_change_history(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村字段级变更历史（时间倒序）"""
    _get_village_or_404(db, village_id, current_user)
    from app.services.audit_enhancement_service import AuditEnhancementService

    history = AuditEnhancementService.get_change_history(db, "supported_village", str(village_id), limit=100)
    return ok_list(history, len(history))


@router.post("/{village_id}/yearly/{year}/validate")
async def validate_yearly_data(
    village_id: int,
    year: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """校验帮扶村年度数据完整性。

    检查项:
    - 8 大板块必填字段
    - 数值合理性（人均收入≥0）
    - 同比变动超 ±50% 预警
        返回错误列表 + 修正建议。
    """
    _get_village_or_404(db, village_id, current_user)
    errors = []
    warnings = []

    # 首次遍历：读取所有板块当年数据并缓存（避免同比比较时重复查询）
    current_year_data = {}
    for section, model in _SECTION_MODEL.items():
        record = db.query(model).filter(
            model.supported_village_id == village_id,
            model.year == year,
        ).first()
        current_year_data[section] = record

        if not record:
            errors.append({
                "section": section,
                "field": None,
                "message": f"板块 [{section}] 未录入数据",
                "suggestion": "请先填写该板块数据",
            })
            continue

        # 通用数值合理性检查
        for col in model.__table__.columns:
            val = getattr(record, col.name, None)
            if isinstance(val, (int, float)) and val < 0:
                errors.append({
                    "section": section,
                    "field": col.name,
                    "message": f"字段 {col.name} 值为 {val}，不能为负数",
                    "suggestion": "请检查原始数据并修正",
                })

    # 同比变动预警（与前一年对比，复用 current_year_data 缓存）
    if year > 0:
        prev_year_data = {}
        for section, model in _SECTION_MODEL.items():
            prev = db.query(model).filter(
                model.supported_village_id == village_id,
                model.year == year - 1,
            ).first()
            if prev:
                for col in model.__table__.columns:
                    prev_year_data[f"{section}.{col.name}"] = getattr(prev, col.name, None)

        if prev_year_data:
            for section, model in _SECTION_MODEL.items():
                record = current_year_data.get(section)
                if not record:
                    continue
                for col in model.__table__.columns:
                    cur = getattr(record, col.name, None)
                    prev_val = prev_year_data.get(f"{section}.{col.name}")
                    if isinstance(cur, (int, float)) and isinstance(prev_val, (int, float)) and prev_val != 0:
                        change_pct = (cur - prev_val) / abs(prev_val) * 100
                        if abs(change_pct) > 50:
                            warnings.append({
                                "section": section,
                                "field": col.name,
                                "message": f"同比变动 {change_pct:+.1f}%（{prev_val} → {cur}），超过 ±50% 阈值",
                                "suggestion": "请核实数据是否准确，如无误可忽略此预警",
                            })

    return success_response(
        data={
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
        }
    )


# ── 区块附件管理 ──


@router.get("/{village_id}/sections/{section}/attachments")
async def get_section_attachments(
    village_id: int,
    section: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村某区块的附件列表"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    attachments = (
        db.query(VillageAttachment)
        .filter(
            VillageAttachment.supported_village_id == village_id,
            VillageAttachment.description.like(f"section:{section}:%"),
        )
        .order_by(VillageAttachment.created_at.desc())
        .all()
    )
    return {
        "code": 200,
        "success": True,
        "data": [
            {
                "id": a.id,
                "fileName": a.file_name,
                "fileSize": a.file_size,
                "fileType": a.mime_type,
                "fileUrl": f"/api/v1/supported-villages/{village_id}/sections/{section}/attachments/{a.id}",
                "createdAt": a.created_at.isoformat() if a.created_at else None,
            }
            for a in attachments
        ],
    }


@router.post("/{village_id}/sections/{section}/attachments")
async def upload_section_attachment(
    village_id: int,
    section: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    import os as _os

    content = await file.read()
    upload_dir = _os.path.join(settings.UPLOAD_DIR, "sections")
    _os.makedirs(upload_dir, exist_ok=True)
    file_path = _os.path.join(upload_dir, f"{village_id}_{section}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(content)

    attachment = VillageAttachment(
        supported_village_id=village_id,
        file_name=file.filename or "unnamed",
        file_path=file_path,
        file_size=len(content),
        mime_type=file.content_type or "application/octet-stream",
        description=f"section:{section}:attachment",
    )
    db.add(attachment)
    safe_commit(db)
    db.refresh(attachment)
    return {
        "code": 200,
        "success": True,
        "data": {"id": attachment.id, "filename": attachment.file_name},
        "message": "上传成功",
    }


@router.get("/{village_id}/sections/{section}/attachments/{attachment_id}")
async def download_section_attachment(
    village_id: int,
    section: str,
    attachment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    import os as _os
    attachment = (
        db.query(VillageAttachment)
        .filter(VillageAttachment.id == attachment_id, VillageAttachment.supported_village_id == village_id)
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")
    if not _os.path.exists(attachment.file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(
        attachment.file_path,
        filename=attachment.file_name,
        media_type=attachment.mime_type or "application/octet-stream",
    )


@router.delete("/{village_id}/sections/{section}/attachments/{attachment_id}")
async def delete_section_attachment(
    village_id: int,
    section: str,
    attachment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    attachment = (
        db.query(VillageAttachment)
        .filter(VillageAttachment.id == attachment_id, VillageAttachment.supported_village_id == village_id)
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")
    db.delete(attachment)
    safe_commit(db)
    return {"code": 200, "success": True, "message": "删除成功"}


# ── 村委数据 ──


@router.post("/{village_id}/committee")
async def save_committee_data(
    village_id: int,
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存帮扶村委数据（旧前端路径兼容入口）。

    根因修复：旧实现查询不带 year（跨年互相覆盖），且对 members 关系属性
    直接 setattr dict 列表导致 flush 时 UnmappedInstanceError → 500。
    现统一委托 _save_section_data：按年隔离 + 成员子表删除重写。
    """
    _get_village_or_404(db, village_id, current_user)
    payload = dict(data)
    year = payload.pop("year", None) or datetime.now().year
    _save_section_data(db, VillageCommitteeInfo, village_id, int(year), payload)
    safe_commit(db)
    await _invalidate_village_cache()
    return {"code": 200, "success": True, "message": "保存成功"}


# ── 区块数据导入 ──

_SECTION_SKIP_ATTRS = frozenset({"id", "supported_village_id", "year", "created_at", "updated_at"})


def _section_label_map(model: Any) -> Dict[str, str]:
    """从模型列注释构建 中文标签 → 属性名 映射（如 "总户数" → "total_households"）。"""
    mapping: Dict[str, str] = {}
    for col in model.__table__.columns:
        if col.name in _SECTION_SKIP_ATTRS:
            continue
        if col.comment:
            # 注释可能带补充说明（如 "总户数" / "进度(%)"），取原文并去掉括号备注
            label = col.comment.split("(")[0].split("（")[0].strip()
            if label:
                mapping.setdefault(label, col.name)
            mapping.setdefault(col.comment.strip(), col.name)
        # 英文属性名本身也可作为表头（兼容机器生成文件）
        mapping.setdefault(col.name, col.name)
    return mapping


def _coerce_section_value(model: Any, attr: str, value: Any) -> Any:
    """按模型列类型把 Excel 单元格值转为合适 Python 类型。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    col = model.__table__.columns.get(attr)
    if col is None:
        return value
    from sqlalchemy import Boolean, Float, Integer, Numeric
    try:
        if isinstance(col.type, Boolean):
            return str(value).strip() in ("是", "1", "True", "true", "yes", "Y")
        if isinstance(col.type, Integer):
            return int(float(str(value).strip()))
        if isinstance(col.type, (Float, Numeric)):
            return float(str(value).strip())
    except (ValueError, TypeError):
        return None
    return value.strip() if isinstance(value, str) else value


def _import_section_sheet(ws, model: Any, village_id: int, year: int, db: Session) -> Dict[str, int]:
    """解析单个工作表并按年写库（upsert）。返回 {imported, failed}。"""
    label_map = _section_label_map(model)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "failed": 0}
    # 表头行：第一个能命中 ≥2 个标签的行（跳过可能的标题/装饰行）
    header_idx = None
    attr_by_col: Dict[int, str] = {}
    for idx, row in enumerate(rows[:10]):
        hits = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).lstrip("*").strip()
            attr = label_map.get(label)
            if attr:
                hits[col_idx] = attr
        if len(hits) >= 2:
            header_idx = idx
            attr_by_col = hits
            break
    if header_idx is None:
        return {"imported": 0, "failed": max(0, len(rows) - 1)}

    imported = 0
    failed = 0
    for row in rows[header_idx + 1:]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        if any(isinstance(c, str) and "示例行" in c for c in row):
            continue
        data: Dict[str, Any] = {}
        for col_idx, attr in attr_by_col.items():
            val = row[col_idx] if col_idx < len(row) else None
            coerced = _coerce_section_value(model, attr, val)
            if coerced is not None:
                data[attr] = coerced
        # 年份列优先取单元格，其次用请求参数
        row_year = data.pop("year", None) or year
        if not data:
            failed += 1
            continue
        try:
            _save_section_data(db, model, village_id, int(row_year), data)
            imported += 1
        except Exception:
            failed += 1
    return {"imported": imported, "failed": failed}


@router.post("/{village_id}/sections/import")
async def import_section_data(
    village_id: int,
    year: Optional[int] = Query(None, description="年度"),
    section_key: Optional[str] = Query(None, description="板块标识"),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导入帮扶村单个区块数据（Excel 表头驱动解析，真实写库）"""
    _get_village_or_404(db, village_id, current_user)
    model = _SECTION_MODEL.get(section_key or "")
    if model is None:
        raise HTTPException(status_code=400, detail=f"未知板块标识: {section_key}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="文件解析失败，请稍后重试或联系管理员")
    target_year = year or datetime.now().year
    result = _import_section_sheet(ws, model, village_id, target_year, db)
    safe_commit(db)
    await _invalidate_village_cache()
    return success_response(
        data={**result, "section_key": section_key, "year": target_year},
        message=f"导入成功 {result['imported']} 行" + (f"，{result['failed']} 行失败" if result["failed"] else ""),
    )


@router.post("/{village_id}/sections/import-all")
async def import_all_sections_data(
    village_id: int,
    year: Optional[int] = Query(None, description="年度"),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导入帮扶村所有区块数据（按工作表名匹配板块，真实写库）"""
    _get_village_or_404(db, village_id, current_user)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
    except Exception:
        raise HTTPException(status_code=400, detail="文件解析失败，请稍后重试或联系管理员")
    # 工作表名 → 板块键：兼容 section_key（population）与模型表名/中文名
    sheet_alias: Dict[str, str] = {}
    for key, model in _SECTION_MODEL.items():
        sheet_alias[key] = key
        sheet_alias[model.__tablename__] = key
        sheet_alias[key.replace("-", "_")] = key
    target_year = year or datetime.now().year
    sections = []
    total_imported = 0
    total_failed = 0
    for name in wb.sheetnames:
        key = sheet_alias.get(name.strip().lower()) or sheet_alias.get(name.strip())
        if key is None:
            continue  # 未识别的工作表跳过，不误报
        result = _import_section_sheet(wb[name], _SECTION_MODEL[key], village_id, target_year, db)
        sections.append({"name": name, "section_key": key, **result})
        total_imported += result["imported"]
        total_failed += result["failed"]
    safe_commit(db)
    await _invalidate_village_cache()
    return success_response(
        data={
            "sheets": len(sections),
            "sections": sections,
            "imported": total_imported,
            "failed": total_failed,
            "year": target_year,
        },
        message=f"导入成功 {total_imported} 行" + (f"，{total_failed} 行失败" if total_failed else ""),
    )


# ── 转移支付资金 ──


class TransitionFundingItem(BaseModel):
    """转移支付资金年度明细项。

    H2b（关键根因修复）：字段名为 camelCase，但 CamelToSnakeMiddleware 会把请求体
    键名 `militaryInvestment` 转为 `military_investment`，导致原字段名无法命中、
    静默回落为 0——这正是"经费保存成功但总额重置为 0"的真正根因。通过
    validation_alias=AliasChoices(camelCase, snake_case) + populate_by_name，同时兼容
    中间件转换后的 snake_case 与未经中间件的 camelCase 两种输入；序列化输出仍以
    字段名(camelCase)为准，保持与 GET /transition-funding 及前端的向后兼容。
    """

    model_config = {"populate_by_name": True}

    year: int
    militaryInvestment: float = Field(
        0, validation_alias=AliasChoices("militaryInvestment", "military_investment")
    )
    localInvestment: float = Field(
        0, validation_alias=AliasChoices("localInvestment", "local_investment")
    )
    totalInvestment: float = Field(
        0, validation_alias=AliasChoices("totalInvestment", "total_investment")
    )
    remarks: Optional[str] = None


class TransitionFundingRequest(BaseModel):
    items: List[TransitionFundingItem]


@router.get("/{village_id}/transition-funding")
async def get_transition_funding(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取转移支付资金按年度明细"""
    village = _get_village_or_404(db, village_id, current_user)
    items = []
    if village.transition_fund_items:
        try:
            items = json.loads(village.transition_fund_items)
        except (json.JSONDecodeError, TypeError):
            items = []
    return success_response(data=items)


@router.post("/{village_id}/transition-funding")
async def save_transition_funding(
    village_id: int,
    data: TransitionFundingRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存转移支付资金数据

    H2（后端兜底，与前端任务#5 语义对齐）：
    - 空 items 代表"用户主动清空经费"的合法意图（前端已通过 fundingLoadFailed
      标志拦截"加载失败导致的异常空提交"），因此允许覆盖为 0，尊重用户意图；
    - 但当检测到"原有非零经费总额被空 items 覆盖为 0"时，调用 write_work_log()
      记录审计日志用于追溯（满足军事合规"写操作留痕"），作为异常/直接 API
      调用场景的兜底防线。
    """
    village = _get_village_or_404(db, village_id, current_user)
    # 记录覆盖前原有总额，用于"非零→零"清空审计判定
    prev_military = float(village.transition_fund_military_total or 0)
    prev_local = float(village.transition_fund_local_total or 0)
    # Sum up military and local totals from the submitted items
    military = sum(item.militaryInvestment for item in data.items)
    local = sum(item.localInvestment for item in data.items)

    # H2：空 items 将把原有非零经费清零——尊重用户主动清空意图，但写审计日志留痕追溯
    if not data.items and (prev_military != 0 or prev_local != 0):
        from app.services.work_log_service import write_work_log

        logger.warning(
            "save_transition_funding: village_id=%s 原有经费(专项=%s, 地方=%s)被空 items 清零，"
            "已按用户主动清空处理并写审计日志",
            village_id, prev_military, prev_local,
        )
        write_work_log(
            db,
            log_type="supported_village",
            action="update",
            entity_id=village_id,
            entity_name=village.village_name,
            user_id=getattr(current_user, "id", None),
            username=getattr(current_user, "username", ""),
            detail=(
                "转移支付资金清空：原专项投入 %.4f 万元、原地方投入 %.4f 万元 → 0"
                % (prev_military, prev_local)
            ),
        )

    village.transition_fund_military_total = military
    village.transition_fund_local_total = local
    # Store per-year breakdown as JSON for retrieval
    items_json = [{
        "year": item.year,
        "militaryInvestment": item.militaryInvestment,
        "localInvestment": item.localInvestment,
        "totalInvestment": item.totalInvestment,
    } for item in data.items]
    village.transition_fund_items = json.dumps(items_json, ensure_ascii=False)
    safe_commit(db)
    await _invalidate_village_cache()
    return ok_list(items_json, len(items_json), message="转移支付资金已保存")


# ── 模板下载 ──


@router.get("/templates/all")
async def download_all_templates(
    current_user: User = Depends(get_current_user),
):
    """下载所有区块模板（Excel 多工作表）"""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sections = ["income", "industry", "infrastructure", "education", "medical", "employment", "population"]
    for s in sections:
        ws = wb.create_sheet(title=s)
        ws.append(["年份", "项目", "数值", "单位", "备注"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=all_templates.xlsx"},
    )
