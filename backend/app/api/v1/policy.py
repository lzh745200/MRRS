"""
政策法规API
支持政策分类、文档管理、导入导出功能
"""

import io
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from starlette.responses import FileResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session

from ...core.cache import cache_manager
from ...core.database import get_db
from ...core.response import ok_list, success_response
from ...core.security import get_current_user
from ...models.policy import Policy, PolicyCategory, PolicyFavorite
from app.core.transaction import safe_commit
from app.api.v1.deps import require_manager_role
from app.services.work_log_service import write_work_log

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/policies", tags=["政策法规"])

# ==================== 辅助函数 ====================


def _level_display_map() -> Dict[str, str]:
    return {
        "national": "国家级",
        "provincial": "省级",
        "municipal": "市级",
        "county": "县级",
        "military": "专项",
        "central_military": "中央部署",
        "theater": "区域",
        "army": "重点",
        "division": "单元",
    }


def _status_display_map() -> Dict[str, str]:
    return {"draft": "草稿", "active": "有效", "invalid": "失效"}


def _safe_isoformat(val) -> Optional[str]:
    """安全地将日期值转为 ISO 格式字符串"""
    if val is None:
        return None
    try:
        if hasattr(val, "isoformat"):
            return val.isoformat()
        return str(val)
    except (TypeError, AttributeError) as e:
        logger.debug(f"日期格式化失败: {e}")
        return str(val) if val else None


def _apply_attachments(policy: Policy, urls) -> None:
    """将前端附件URL列表映射到政策主文件（首附件作为预览/下载主文件）"""
    if not urls:
        return
    clean = [u for u in urls if isinstance(u, str) and u.strip()]
    if not clean:
        return
    import os as _os

    from app.core.config import settings

    first = clean[0]
    # /uploads/xxx → 本地绝对路径
    if first.startswith("/uploads/"):
        rel = first[len("/uploads/"):].replace("/", _os.sep)
        local = _os.path.join(_os.path.abspath(settings.UPLOAD_DIR), rel)
    else:
        local = first
    policy.file_path = local
    ext = _os.path.splitext(local)[1].lower().lstrip(".")
    policy.file_type = ext or None
    try:
        policy.file_size = _os.path.getsize(local) if _os.path.exists(local) else 0
    except OSError:
        policy.file_size = 0


def _policy_to_frontend(policy: Policy) -> Dict[str, Any]:
    """将数据库 Policy 对象转换为前端期望的格式"""
    level_val = str(policy.level) if policy.level else ""
    status_val = str(policy.status) if policy.status else "draft"
    category_val = str(policy.category) if policy.category else ""

    # 前端需要的 category_name / level_name / status_name
    level_names = _level_display_map()
    status_names = _status_display_map()

    return {
        "id": policy.id,
        "title": policy.title,
        "code": policy.code,
        "content": policy.content or "",
        "summary": policy.summary,
        "keywords": policy.keywords,
        # 分类 & 层级
        "category": category_val,
        "category_name": (
            "专项政策" if category_val == "military" else "地方政策" if category_val == "local" else category_val
        ),
        "organization_level": level_val,
        "level": level_val,
        "level_name": level_names.get(level_val, level_val),
        # 状态
        "status": status_val,
        "status_name": status_names.get(status_val, status_val),
        # 日期
        "publish_date": _safe_isoformat(policy.issue_date) or _safe_isoformat(policy.created_at),
        "issue_date": _safe_isoformat(policy.issue_date),
        "effective_date": _safe_isoformat(policy.effective_date),
        # 发布信息
        "issuing_authority": policy.issuing_authority,
        "department": policy.issuing_authority,  # 前端使用 department 显示
        "document_number": policy.code,
        # 附件（URL 形式，供前端展示/下载）
        "attachment_urls": _attachment_urls_of(policy),
        # 统计
        "view_count": policy.view_count or 0,
        "download_count": policy.download_count or 0,
        # 时间
        "created_at": _safe_isoformat(policy.created_at),
        "updated_at": _safe_isoformat(policy.updated_at),
    }


# ==================== Pydantic模型 ====================


class PolicyCategoryBase(BaseModel):
    name: str
    code: Optional[str] = None
    parent_id: Optional[int] = None
    description: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True


class PolicyCategoryCreate(PolicyCategoryBase):
    pass


class PolicyCategoryResponse(PolicyCategoryBase):
    id: int
    created_at: Optional[datetime] = None

    model_config = ConfigDict(from_attributes=True)


class PolicyCreateRequest(BaseModel):
    """创建/更新政策请求（兼容前端字段）"""

    title: str
    content: Optional[str] = None
    category: Optional[str] = None
    organization_level: Optional[str] = None
    level: Optional[str] = None
    status: Optional[str] = "draft"
    publish_date: Optional[str] = None
    effective_date: Optional[str] = None
    expiry_date: Optional[str] = None
    issuing_authority: Optional[str] = None
    document_number: Optional[str] = None
    code: Optional[str] = None
    summary: Optional[str] = None
    keywords: Optional[str] = None
    attachment_urls: Optional[List[str]] = None


class PolicyUpdateRequest(BaseModel):
    """更新政策请求"""

    title: Optional[str] = None
    content: Optional[str] = None
    category: Optional[str] = None
    organization_level: Optional[str] = None
    level: Optional[str] = None
    status: Optional[str] = None
    publish_date: Optional[str] = None
    effective_date: Optional[str] = None
    expiry_date: Optional[str] = None
    issuing_authority: Optional[str] = None
    document_number: Optional[str] = None
    code: Optional[str] = None
    summary: Optional[str] = None
    keywords: Optional[str] = None
    attachment_urls: Optional[List[str]] = None


# ==================== 政策分类API ====================


@router.get("/categories")
async def get_categories(
    parent_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
    response: Any = None,
):
    """获取政策分类列表 —— 兼容前端两种调用方式"""
    try:
        query = db.query(PolicyCategory)
        if parent_id is not None:
            query = query.filter(PolicyCategory.parent_id == parent_id)
        if is_active is not None:
            query = query.filter(PolicyCategory.is_active == is_active)
        cats = query.order_by(PolicyCategory.sort_order, PolicyCategory.id).all()
        # 如果有分类数据，按后端结构返回
        if cats:
            return [
                {
                    "id": c.id,
                    "name": c.name,
                    "code": c.code,
                    "parent_id": c.parent_id,
                    "description": c.description,
                    "sort_order": c.sort_order,
                    "is_active": c.is_active,
                    "created_at": c.created_at.isoformat() if c.created_at else None,
                }
                for c in cats
            ]
    except (ValueError, TypeError, AttributeError) as e:
        logger.warning(f"查询分类表失败: {e}")

    # 返回前端期望的静态分类配置
    return success_response(data={
        "military": {
            "label": "专项政策",
            "levels": [
                {"value": "central_military", "label": "中央部署"},
                {"value": "theater", "label": "区域"},
                {"value": "army", "label": "重点"},
                {"value": "division", "label": "单元"},
            ],
        },
        "local": {
            "label": "地方政策",
            "levels": [
                {"value": "national", "label": "国家级"},
                {"value": "provincial", "label": "省级"},
                {"value": "municipal", "label": "市级"},
                {"value": "county", "label": "县级"},
            ],
        },
    })


@router.get("/categories/tree")
async def get_category_tree(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """获取政策分类树形结构"""
    try:
        categories = (
            db.query(PolicyCategory)
            .filter(PolicyCategory.is_active == True)  # noqa: E712
            .order_by(PolicyCategory.sort_order, PolicyCategory.id)
            .all()
        )

        cat_map: Dict[int, Dict[str, Any]] = {
            cat.id: {
                "id": cat.id,
                "name": cat.name,
                "code": cat.code,
                "parent_id": cat.parent_id,
                "children": [],
            }
            for cat in categories
        }

        tree: List[Dict[str, Any]] = []
        for cat in categories:
            node = cat_map[cat.id]
            if cat.parent_id and cat.parent_id in cat_map:
                cat_map[cat.parent_id]["children"].append(node)
            else:
                tree.append(node)
        return tree
    except (ValueError, TypeError, KeyError) as e:
        logger.warning(f"获取分类树失败: {e}")
        return []


@router.post("/categories")
async def create_category(
    data: PolicyCategoryCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建政策分类"""
    require_manager_role(current_user)
    if data.code:
        existing = db.query(PolicyCategory).filter(PolicyCategory.code == data.code).first()
        if existing:
            raise HTTPException(status_code=400, detail="分类编码已存在")

    category = PolicyCategory(**data.model_dump())
    db.add(category)
    safe_commit(db)
    db.refresh(category)
    return success_response(
        data={"id": category.id, "name": category.name, "code": category.code},
        message="成功",
    )


@router.put("/categories/{category_id}")
async def update_category(
    category_id: int,
    data: PolicyCategoryCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新政策分类"""
    require_manager_role(current_user)
    category = db.query(PolicyCategory).filter(PolicyCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(category, key, value)
    safe_commit(db)
    db.refresh(category)
    return success_response(
        data={"id": category.id, "name": category.name, "code": category.code},
        message="成功",
    )


@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除政策分类"""
    require_manager_role(current_user)
    category = db.query(PolicyCategory).filter(PolicyCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")

    children = db.query(PolicyCategory).filter(
        PolicyCategory.parent_id == category_id,
        PolicyCategory.is_active == True  # noqa: E712 -- SQLAlchemy boolean filter
    ).count()
    if children > 0:
        raise HTTPException(status_code=400, detail="该分类下有子分类，无法删除")

    db.delete(category)
    safe_commit(db)
    return success_response(message="删除成功")


# ==================== 统计 API ====================


@router.get("/statistics")
async def get_policy_statistics(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """获取政策统计数据（前端 Category.vue 使用）

    优化：使用数据库聚合查询替代 Python 循环，避免 N+1 查询问题
    """
    from sqlalchemy import func

    # 使用数据库聚合查询，一次性获取所有统计数据
    stats = (
        db.query(Policy.category, Policy.level, func.count(Policy.id).label("count"))
        .group_by(Policy.category, Policy.level)
        .all()
    )

    military_levels: Dict[str, int] = {}
    local_levels: Dict[str, int] = {}
    military_total = 0
    local_total = 0

    for stat in stats:
        cat = str(stat.category) if stat.category else ""
        lvl = str(stat.level) if stat.level else "unknown"
        count = int(stat.count)

        if cat == "military":
            military_total += count
            military_levels[lvl] = military_levels.get(lvl, 0) + count
        else:
            local_total += count
            local_levels[lvl] = local_levels.get(lvl, 0) + count

    return success_response(
        data={
            "military": {"total": military_total, "levels": military_levels},
            "local": {"total": local_total, "levels": local_levels},
        }
    )


# ==================== 导入导出API ====================


@router.get("/import/template")
async def download_import_template():
    """下载政策导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_policy_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''policy_import_template.xlsx"},
    )


@router.post("/import")
async def import_policies(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    从 Excel 导入政策
    前端调用 POST /policies/import
    返回格式: { imported: int, errors: list }
    """
    require_manager_role(current_user)
    from ...services.policy_import_service import import_policies_from_excel
    return await import_policies_from_excel(file, db, current_user)


# 保留旧路径做兼容


@router.post("/import/excel")
async def import_policies_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入政策（旧路径兼容）"""
    require_manager_role(current_user)
    return await import_policies(file=file, current_user=current_user, db=db)


def _build_export_workbook(policies_list: List[Policy]):
    """构建导出 Excel 工作簿（PDF/WPS/Excel 共用）"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    level_display = _level_display_map()
    status_display = _status_display_map()

    wb = Workbook()
    ws = wb.active
    ws.title = "政策法规"

    headers = [
        "序号",
        "政策标题",
        "政策文号",
        "分类",
        "政策级别",
        "发布机关",
        "发布日期",
        "生效日期",
        "状态",
        "关键词",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, policy in enumerate(policies_list, 1):
        level_val = str(policy.level) if policy.level else ""
        status_val = str(policy.status) if policy.status else "draft"
        cat = str(policy.category) if policy.category else ""
        ws.append(
            [
                idx,
                policy.title,
                policy.code or "",
                ("专项政策" if cat == "military" else "地方政策" if cat == "local" else cat),
                level_display.get(level_val, level_val),
                policy.issuing_authority or "",
                policy.issue_date.strftime("%Y-%m-%d") if policy.issue_date else "",
                (policy.effective_date.strftime("%Y-%m-%d") if policy.effective_date else ""),
                status_display.get(status_val, status_val),
                policy.keywords or "",
            ]
        )

    widths = {
        "A": 8,
        "B": 50,
        "C": 20,
        "D": 12,
        "E": 12,
        "F": 20,
        "G": 15,
        "H": 15,
        "I": 12,
        "J": 30,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return wb


def _query_policies_for_export(db: Session, params: dict) -> List[Policy]:
    """根据筛选条件查询要导出的政策"""
    query = db.query(Policy)
    if params.get("category"):
        query = query.filter(Policy.category == params["category"])
    if params.get("organization_level"):
        query = query.filter(Policy.level == params["organization_level"])
    if params.get("status"):
        query = query.filter(Policy.status == params["status"])
    if params.get("search"):
        kw = params["search"]
        query = query.filter((Policy.title.contains(kw)) | (Policy.keywords.contains(kw)))
    return query.order_by(Policy.created_at.desc()).all()


def _build_policies_pdf(policies_list: List[Policy]) -> bytes:
    """构建政策列表 PDF（reportlab + Adobe CID 中文字体，离线可用）"""
    from datetime import datetime as _dt

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        _FONT = "STSong-Light"
    except Exception:
        _FONT = "Helvetica"

    level_display = _level_display_map()
    status_display = _status_display_map()

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle", parent=styles["Title"], fontName=_FONT, fontSize=16,
        alignment=TA_CENTER, spaceAfter=6,
    )
    head_style = ParagraphStyle(
        "PdfHead", parent=styles["Normal"], fontName=_FONT, fontSize=11,
        alignment=TA_CENTER, spaceAfter=10,
    )
    foot_style = ParagraphStyle(
        "PdfFoot", parent=styles["Normal"], fontName=_FONT, fontSize=8, textColor=colors.grey,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title="帮扶政策清单",
    )
    story = []
    story.append(Paragraph("帮扶政策清单", title_style))
    story.append(Paragraph(
        f"导出时间：{_dt.now().strftime('%Y-%m-%d %H:%M')}    共 {len(policies_list)} 条",
        head_style,
    ))
    story.append(Spacer(1, 6))

    headers = ["序号", "政策标题", "政策文号", "分类", "级别", "发布机关", "发布日期", "状态", "关键词"]
    data = [headers]
    for idx, p in enumerate(policies_list, 1):
        level_val = str(p.level) if p.level else ""
        status_val = str(p.status) if p.status else "draft"
        cat = str(p.category) if p.category else ""
        data.append([
            str(idx),
            str(p.title or ""),
            str(p.code or ""),
            ("专项政策" if cat == "military" else "地方政策" if cat == "local" else cat),
            level_display.get(level_val, level_val),
            str(p.issuing_authority or ""),
            p.issue_date.strftime("%Y-%m-%d") if p.issue_date else "",
            status_display.get(status_val, status_val),
            str(p.keywords or ""),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E6B55")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B8C6BE")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7F4")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph("— 帮扶管理信息系统自动导出 —", foot_style))

    doc.build(story)
    return buffer.getvalue()


@router.get("/export/excel")
async def export_policies_excel(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策到 Excel"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    wb = _build_export_workbook(policies_list)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=policies.xlsx"},
    )


@router.get("/export/pdf")
async def export_policies_pdf(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策列表为真实 PDF 文档"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    pdf_bytes = _build_policies_pdf(policies_list)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=policies_export.pdf"},
    )


@router.get("/export/wps")
async def export_policies_wps(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策列表为 WPS 兼容表格（xlsx，WPS/Excel 均可打开）"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    wb = _build_export_workbook(policies_list)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=policies_export_wps.xlsx"},
    )


@router.get("/types")
async def get_policy_types(db: Session = Depends(get_db)):
    """获取政策类型选项 — 合并预定义类型与数据库中的实际分类"""
    # 预定义政策类型
    types = [
        {"value": "military", "label": "专项政策"},
        {"value": "local", "label": "地方政策"},
        {"value": "national", "label": "国家政策"},
        {"value": "provincial", "label": "省级政策"},
        {"value": "municipal", "label": "市级政策"},
        {"value": "county", "label": "县级政策"},
        {"value": "other", "label": "其他"},
    ]

    # 尝试从数据库中获取已有的政策分类（如果 PolicyCategory 模型已启用）
    try:
        categories = db.query(PolicyCategory).filter(PolicyCategory.is_active == True).all()  # noqa: E712
        if categories:
            db_types = [{"value": c.code or c.name, "label": c.name} for c in categories]
            # 合并去重：数据库分类优先
            existing_codes = {t["value"] for t in types}
            for dt in db_types:
                if dt["value"] not in existing_codes:
                    types.append(dt)
    except Exception:
        logger.debug("获取政策分类时出错，使用预定义类型")

    return success_response(data=types)


@router.get("/options/levels")
async def get_level_options():
    """获取政策级别选项"""
    return [
        {"value": "national", "label": "国家级"},
        {"value": "provincial", "label": "省级"},
        {"value": "municipal", "label": "市级"},
        {"value": "county", "label": "县级"},
        {"value": "military", "label": "专项"},
    ]


@router.get("/options/statuses")
async def get_status_options():
    """获取政策状态选项"""
    return [
        {"value": "active", "label": "有效"},
        {"value": "invalid", "label": "失效"},
        {"value": "draft", "label": "草稿"},
    ]


# ==================== 文件上传与预览 API ====================


@router.post("/{policy_id}/upload")
async def upload_policy_file(
    policy_id: int,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传政策附件文件（支持 pdf/doc/docx/pptx）"""
    require_manager_role(current_user)
    import os

    from app.core.config import settings

    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    # 校验文件类型
    ext = os.path.splitext(file.filename or "")[1].lower().lstrip(".")
    allowed = {"pdf", "doc", "docx", "pptx"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型，仅支持: {', '.join(allowed)}")

    content = await file.read()
    # 限制 50MB
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小超过50MB限制")

    # 存储到 uploads/policies/ 目录
    upload_dir = os.path.join(settings.UPLOAD_DIR, "policies")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"policy_{policy_id}_{int(datetime.now().timestamp())}.{ext}"
    file_path = os.path.join(upload_dir, safe_name)

    with open(file_path, "wb") as f:
        f.write(content)

    # 更新数据库
    setattr(policy, "file_path", file_path)
    setattr(policy, "file_size", len(content))
    setattr(policy, "file_type", ext)
    safe_commit(db)

    return success_response(
        message="上传成功",
        data={
            "file_path": file_path,
            "file_size": len(content),
            "file_type": ext,
        },
    )


@router.get("/{policy_id}/preview")
async def preview_policy_file(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """预览政策附件文件（返回文件流或HTML）"""
    import os

    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    if not policy.file_path or not os.path.exists(policy.file_path):
        # 没有附件，返回正文内容作HTML预览
        html = f"<html><body><h1>{policy.title}</h1><div>{policy.content or '无内容'}</div></body></html>"
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html",
        )

    ext = (policy.file_type or "").lower()

    if ext == "pdf":
        return FileResponse(
            path=policy.file_path,
            media_type="application/pdf",
            filename=os.path.basename(policy.file_path),
        )
    elif ext in ("doc", "docx"):
        # 使用 mammoth 将 docx 转换为 HTML
        try:
            import mammoth

            with open(policy.file_path, "rb") as f:
                result = mammoth.convert_to_html(f)
            html_style = "body{font-family:SimSun,serif;padding:20px;max-width:800px;margin:0 auto}"
            html = (
                f"<html><head><meta charset='utf-8'><style>{html_style}</style></head>"
                f"<body><h2>{policy.title}</h2>{result.value}</body></html>"
            )
            return StreamingResponse(
                io.BytesIO(html.encode("utf-8")),
                media_type="text/html",
            )
        except ImportError:  # pragma: no cover
            # mammoth 未安装，返回下载
            return FileResponse(
                path=policy.file_path,
                media_type="application/octet-stream",
                filename=os.path.basename(policy.file_path),
            )
    else:
        # 其他类型直接下载
        return FileResponse(
            path=policy.file_path,
            media_type="application/octet-stream",
            filename=os.path.basename(policy.file_path),
        )


@router.get("/{policy_id}/download")
async def download_policy_file(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载政策附件"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")
    if not policy.file_path or not os.path.exists(policy.file_path):
        raise HTTPException(status_code=404, detail="附件文件不存在")

    current_count = policy.download_count or 0
    setattr(policy, "download_count", current_count + 1)
    safe_commit(db)

    return FileResponse(
        path=policy.file_path,
        media_type="application/octet-stream",
        filename=os.path.basename(policy.file_path),
    )


# ==================== 批量操作 API ====================


@router.post("/batch-delete")
async def batch_delete_policies(data: dict, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """批量删除政策"""
    require_manager_role(current_user)
    raw_ids = data.get("ids", [])
    if not raw_ids:
        raise HTTPException(status_code=400, detail="未提供要删除的ID")
    if not isinstance(raw_ids, list):
        raise HTTPException(status_code=422, detail="ids 必须是数组")
    ids: List[int] = []
    for rid in raw_ids:
        if isinstance(rid, bool) or not isinstance(rid, int):
            raise HTTPException(status_code=422, detail="ids 只能包含正整数")
        if rid <= 0:
            raise HTTPException(status_code=422, detail="ids 只能包含正整数")
        ids.append(rid)
    if len(ids) > 1000:
        raise HTTPException(status_code=422, detail="单次最多删除1000条")

    deleted = db.query(Policy).filter(Policy.id.in_(ids)).delete(synchronize_session=False)
    safe_commit(db)
    await cache_manager.delete("policies:list")
    return success_response(message=f"成功删除{deleted}条政策", data={"deleted": deleted})


# ==================== 政策CRUD API ====================


@router.get("")
async def get_policies(
    # 前端参数 (store 发送)
    skip: Optional[int] = Query(None, description="偏移量"),
    limit: Optional[int] = Query(None, description="数量限制"),
    category: Optional[str] = Query(None, description="分类: military/local"),
    organization_level: Optional[str] = Query(None, description="组织层级"),
    search: Optional[str] = Query(None, description="搜索关键字"),
    order_by: Optional[str] = Query(None, description="排序字段"),
    order_desc: Optional[bool] = Query(None, description="是否降序"),
    # 年度/文号筛选
    year: Optional[int] = Query(None, description="发布年度筛选"),
    document_code: Optional[str] = Query(None, description="文号筛选"),
    # 兼容旧参数
    page: Optional[int] = Query(None, ge=1),
    page_size: Optional[int] = Query(None, ge=1, le=200),
    keyword: Optional[str] = Query(None),
    level: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取政策列表 —— 兼容前端 skip/limit 和旧 page/page_size 参数"""
    _cache_key = "policies:list"
    _no_filter = not any([category, organization_level, search, order_by, year, document_code, keyword, level, status])
    _is_default_page = skip in (None, 0) and limit in (None, 20) and page is None
    if _no_filter and _is_default_page:
        cached = await cache_manager.get(_cache_key)
        if cached is not None:
            return cached

    query = db.query(Policy).filter(Policy.is_active == True)  # noqa: E712

    # 关键字搜索
    kw = search or keyword
    if kw:
        query = query.filter((Policy.title.contains(kw)) | (Policy.code.contains(kw)) | (Policy.keywords.contains(kw)))
    # 分类过滤
    if category:
        query = query.filter(Policy.category == category)
    # 层级过滤
    lvl = organization_level or level
    if lvl:
        query = query.filter(Policy.level == lvl)
    # 状态过滤
    if status:
        query = query.filter(Policy.status == status)
    # 年度筛选
    if year:
        from sqlalchemy import extract

        query = query.filter(extract("year", Policy.issue_date) == year)
    # 文号筛选
    if document_code:
        query = query.filter(Policy.code.contains(document_code))

    total = query.count()

    # 排序

    sort_col: Any = Policy.created_at
    if order_by == "publish_date":
        sort_col = Policy.issue_date
    elif order_by == "title":
        sort_col = Policy.title
    if order_desc is False:
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    # 分页
    if skip is not None:
        offset = max(skip, 0)
        lim = limit or 10
    elif page is not None:
        offset = (page - 1) * (page_size or 20)
        lim = page_size or 20
    else:
        offset = 0
        lim = limit or 20

    items = query.offset(offset).limit(lim).all()

    items_list = [_policy_to_frontend(p) for p in items]
    result = ok_list(items=items_list, total=total, page=(offset // lim) + 1 if lim else 1, page_size=lim)
    if _no_filter and _is_default_page:
        await cache_manager.set(_cache_key, result, ttl=300)
    return result


@router.get("/{policy_id}/related")
async def get_related_policies(
    policy_id: int,
    limit: int = Query(5, ge=1, le=20),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取相关政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    query = db.query(Policy).filter(Policy.id != policy_id)
    if policy.category:
        query = query.filter(Policy.category == policy.category)
    related = query.order_by(Policy.created_at.desc()).limit(limit).all()
    return [_policy_to_frontend(p) for p in related]


@router.get("/search")
async def search_policies(
    q: str = Query("", description="搜索关键词"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """全文检索帮扶政策（FTS5 + 关键词高亮）"""
    from app.services.policy_fts_service import search_policies_fts
    results = search_policies_fts(db, q, limit=limit, offset=offset)
    return ok_list(results, len(results), query=q)


@router.get("/{policy_id}")
async def get_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取政策详情"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    current_count = policy.view_count or 0
    setattr(policy, "view_count", current_count + 1)
    safe_commit(db)

    return success_response(data=_policy_to_frontend(policy))


@router.post("")
async def create_policy(  # noqa: C901
    data: PolicyCreateRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建政策"""
    require_manager_role(current_user)
    try:
        issue_date = None
        effective_date = None
        if data.publish_date:
            try:
                issue_date = datetime.fromisoformat(data.publish_date)
            except (ValueError, TypeError):
                logger.warning("Invalid publish_date format: %s", data.publish_date)
        if data.effective_date:
            try:
                effective_date = datetime.fromisoformat(data.effective_date)
            except (ValueError, TypeError):
                logger.warning("Invalid effective_date format: %s", data.effective_date)

        policy = Policy(
            title=data.title,
            content=data.content or "",
            category=data.category or "local",
            level=data.organization_level or data.level or "national",
            status=data.status or "draft",
            issuing_authority=data.issuing_authority,
            code=data.document_number or data.code,
            issue_date=issue_date,
            effective_date=effective_date,
            summary=data.summary,
            keywords=data.keywords,
            created_by=current_user.id,
            organization_id=getattr(current_user, "organization_id", None),
        )
        if data.attachment_urls:
            _apply_attachments(policy, data.attachment_urls)
        db.add(policy)
        safe_commit(db)
        db.refresh(policy)
        await cache_manager.delete("policies:list")
        # FTS 索引同步（先确保 FTS 表存在，避免首条记录静默丢失索引）
        from app.services.policy_fts_service import ensure_fts_table, sync_policy_to_fts
        try:
            ensure_fts_table(db)
        except Exception:
            logger.warning("FTS 表初始化失败", exc_info=True)
        try:
            sync_policy_to_fts(db, policy.id)
        except Exception:
            logger.warning("政策 FTS 索引同步失败（不影响主流程）", exc_info=True)
        # 审计日志
        try:
            write_work_log(db, "policy", "create", policy.id, f"创建政策: {policy.title}",
                           user_id=current_user.id, username=getattr(current_user, "username", ""))
        except Exception:
            logger.debug("记录工作日志失败", exc_info=True)
        return success_response(data=_policy_to_frontend(policy))
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建政策失败: {e}")
        raise HTTPException(status_code=500, detail=f"创建政策失败: {str(e)}")


def _coerce_date_fields(update_data: dict) -> None:
    """将字符串日期字段转换为 datetime; 非法值移除"""
    for date_field in ["issue_date", "effective_date"]:
        if date_field in update_data and isinstance(update_data[date_field], str):
            try:
                update_data[date_field] = datetime.fromisoformat(update_data[date_field])
            except (ValueError, TypeError):
                update_data.pop(date_field)


def _filter_valid_columns(policy_cls, update_data: dict):
    """剔除模型不存在的字段; 返回 attachment_urls"""
    attachment_urls = update_data.pop("attachment_urls", None)
    valid_columns = {col.name for col in policy_cls.__table__.columns}
    for key in list(update_data.keys()):
        if key not in valid_columns:
            update_data.pop(key)
    return attachment_urls


@router.put("/{policy_id}")
async def update_policy(
    policy_id: int,
    data: PolicyUpdateRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新政策"""
    require_manager_role(current_user)
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    try:
        update_data = data.model_dump(exclude_unset=True)

        field_mapping = {
            "organization_level": "level",
            "document_number": "code",
            "publish_date": "issue_date",
        }

        for frontend_key, db_key in field_mapping.items():
            if frontend_key in update_data:
                update_data[db_key] = update_data.pop(frontend_key)

        # 处理日期字段
        _coerce_date_fields(update_data)

        # 移除模型不存在的字段（如 URL 预处理的 attachment_urls）
        attachment_urls = _filter_valid_columns(Policy, update_data)
        if attachment_urls is not None:
            _apply_attachments(policy, attachment_urls)

        for key, value in update_data.items():
            setattr(policy, key, value)

        safe_commit(db)
        db.refresh(policy)
        await cache_manager.delete("policies:list")
        # FTS 索引同步（先确保 FTS 表存在，避免首条记录静默丢失索引）
        from app.services.policy_fts_service import ensure_fts_table, sync_policy_to_fts
        try:
            ensure_fts_table(db)
        except Exception:
            logger.warning("FTS 表初始化失败", exc_info=True)
        try:
            sync_policy_to_fts(db, policy.id)
        except Exception:
            logger.warning("政策 FTS 索引同步失败（不影响主流程）", exc_info=True)
        # 审计日志
        try:
            write_work_log(db, "policy", "update", policy.id, f"更新政策: {policy.title}",
                           user_id=current_user.id, username=getattr(current_user, "username", ""))
        except Exception:
            logger.debug("记录工作日志失败", exc_info=True)
        return success_response(data=_policy_to_frontend(policy))
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新政策失败 (id={policy_id}): {e}")
        raise HTTPException(status_code=500, detail=f"更新政策失败: {str(e)}")


@router.delete("/{policy_id}")
async def delete_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除政策（软删除）"""
    require_manager_role(current_user)
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    policy.is_active = False
    safe_commit(db)
    await cache_manager.delete("policies:list")
    # FTS 索引同步
    from app.services.policy_fts_service import remove_policy_from_fts
    remove_policy_from_fts(db, policy_id)
    # 审计日志
    try:
        write_work_log(db, "policy", "delete", policy_id, f"删除政策: {policy.title}",
                       user_id=current_user.id, username=getattr(current_user, "username", ""))
    except Exception:
        logger.debug("记录工作日志失败", exc_info=True)
    return success_response(message="删除成功")


@router.post("/{policy_id}/publish")
async def publish_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """发布政策"""
    require_manager_role(current_user)
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    setattr(policy, "status", "active")
    safe_commit(db)
    await cache_manager.delete("policies:list")
    return success_response(message="发布成功")


@router.post("/{policy_id}/archive")
async def archive_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """归档政策"""
    require_manager_role(current_user)
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    setattr(policy, "status", "invalid")
    safe_commit(db)
    await cache_manager.delete("policies:list")
    return success_response(message="归档成功")


# ==================== 收藏API ====================


@router.post("/{policy_id}/favorite")
async def add_favorite(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """收藏政策（仅能为当前登录用户收藏，忽略客户端传入的 user_id 以防 IDOR）"""
    user_id = current_user.id
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    existing = (
        db.query(PolicyFavorite)
        .filter(PolicyFavorite.policy_id == policy_id, PolicyFavorite.user_id == user_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="已收藏该政策")

    favorite = PolicyFavorite(policy_id=policy_id, user_id=user_id)
    db.add(favorite)
    safe_commit(db)
    return success_response(message="收藏成功")


@router.delete("/{policy_id}/favorite")
async def remove_favorite(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """取消收藏（仅能取消当前登录用户的收藏，忽略客户端传入的 user_id 以防 IDOR）"""
    user_id = current_user.id
    favorite = (
        db.query(PolicyFavorite)
        .filter(PolicyFavorite.policy_id == policy_id, PolicyFavorite.user_id == user_id)
        .first()
    )
    if not favorite:
        raise HTTPException(status_code=404, detail="未收藏该政策")

    db.delete(favorite)
    safe_commit(db)
    return success_response(message="取消收藏成功")


@router.get("/user/{user_id}/favorites")
async def get_user_favorites(
    user_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取用户收藏的政策（仅允许查看自己的收藏，防止越权读取他人数据）"""
    if user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能查看自己的收藏")
    favorites = db.query(PolicyFavorite).filter(PolicyFavorite.user_id == user_id).all()
    policy_ids = [f.policy_id for f in favorites]
    if not policy_ids:
        return []
    items = db.query(Policy).filter(Policy.id.in_(policy_ids)).all()
    return [_policy_to_frontend(p) for p in items]


def _attachment_urls_of(policy: Policy) -> list:
    """政策附件以 URL 形式输出（/uploads/xxx），供前端展示与下载"""
    fp = getattr(policy, "file_path", None)
    if not fp:
        return []
    import os as _os

    from app.core.config import settings

    normalized = _os.path.normpath(_os.path.abspath(fp))
    base = _os.path.normpath(_os.path.abspath(settings.UPLOAD_DIR))
    if normalized.startswith(base):
        rel = _os.path.relpath(normalized, base).replace(_os.sep, "/")
        return [f"/uploads/{rel}"]
    return [fp]
