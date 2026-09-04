"""Policy import from Excel — extracted from api/v1/policy.py to reduce C901 complexity."""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.policy import Policy
from app.core.upload_security import validate_excel_upload
from app.core.logging import logger
from app.core.transaction import safe_commit


async def import_policies_from_excel(
    file: UploadFile,
    db: Session,
    current_user=None,
) -> Dict[str, Any]:
    """从 Excel 导入政策

    返回格式: { imported: int, errors: list, errorRows: list, total: int }
    """
    if not validate_excel_upload(file):
        raise HTTPException(status_code=400, detail="文件校验失败: 不是有效的 Excel 文件")
    content = await file.read()

    try:
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        level_map = {
            "国家级": "national",
            "省级": "provincial",
            "市级": "municipal",
            "县级": "county",
            "专项": "military",
            "中央部署": "central_military",
            "区域": "theater",
            "重点": "army",
            "单元": "division",
        }
        status_map = {
            "草稿": "draft",
            "有效": "active",
            "失效": "invalid",
            "已发布": "active",
            "已归档": "invalid",
            "已过期": "invalid",
            # 与导入模板状态列说明一致（excel_template_service.POLICY_FIELDS）
            "现行有效": "active",
            "已修订": "active",
            "即将实施": "draft",
            "已废止": "invalid",
        }

        # 模板布局：标题区(1-5行) + 表头行 + 示例行 + 数据行；
        # 同时兼容用户自制"第1行表头"文件 → 自动探测表头行
        header_row = _find_header_row(ws, "政策标题")
        data_start = header_row + 1

        imported = 0
        errors: List[Dict[str, Any]] = []
        error_rows: List[int] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), start=data_start
        ):
            if not row or len(row) < 2:
                continue
            # 跳过模板示例行（行尾合并单元格标注"← 示例行（导入时自动跳过）"）与说明/页脚行
            if any(
                isinstance(c, str) and ("示例行" in c or "填写说明" in c or "帮扶管理信息系统 v" in c)
                for c in row
            ):
                continue
            if not row[1]:
                # 整行为空（预格式化空行/分隔区）→ 静默跳过，不误报错误
                if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                    continue
                errors.append({"row": row_idx, "title": "", "error": "政策标题不能为空"})
                error_rows.append(row_idx)
                continue

            try:
                result = _process_policy_row(row, row_idx, level_map, status_map, db)
                if result.get("error"):
                    errors.append(result)
                    error_rows.append(row_idx)
                elif result.get("duplicate"):
                    errors.append(result)
                    error_rows.append(row_idx)
                else:
                    imported += 1
            except (ValueError, TypeError, KeyError) as e:
                errors.append(_make_row_error(row, row_idx, str(e)))
                error_rows.append(row_idx)

        safe_commit(db)

        return {
            "imported": imported,
            "errors": errors,
            "errorRows": error_rows,
            "total": imported + len(errors),
        }
    except Exception as e:
        db.rollback()
        # 完整栈仅进服务端日志；detail 不得内插异常原文（W1 不变量 #6）
        logger.error(f"导入政策失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="导入失败，请稍后重试或联系管理员")


def _find_header_row(ws, required_label: str, max_scan: int = 10) -> int:
    """在前 max_scan 行内探测表头行（含"序号"与 required_label 列标签）。

    官方模板表头在第 6 行（前 5 行为装饰标题区）；
    用户自制文件通常第 1 行即表头。探测失败时回退为 1（保持旧行为）。
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).lstrip("*").strip() for c in row if c is not None]
        if required_label in cells and "序号" in cells:
            return idx
    return 1


def _process_policy_row(row, row_idx, level_map, status_map, db) -> Dict[str, Any]:
    """处理单行 Excel 数据，返回结果字典。"""
    issue_date = _parse_date_cell(row, 5, row_idx, "发布日期")
    effective_date = _parse_date_cell(row, 6, row_idx, "生效日期")

    level_str = _safe_str(row, 3)
    level_val = level_map.get(level_str, "national") if level_str else "national"

    status_str = _safe_str(row, 7)
    status_val = status_map.get(status_str, "active") if status_str else "active"

    category_val = "military" if level_val == "military" else "local"
    content_val = _safe_str(row, 9, default="")
    title_val = _safe_str(row, 1, default=f"未命名政策{row_idx}")
    code_val = _safe_str(row, 2)
    authority_val = _safe_str(row, 4)
    keywords_val = _safe_str(row, 8)

    if code_val:
        existing = db.query(Policy).filter(Policy.code == code_val).first()
        if existing:
            return {"row": row_idx, "title": title_val, "error": f"文号\u201c{code_val}\u201d已存在，已跳过"}

    policy = Policy(
        title=title_val,
        code=code_val,
        level=level_val,
        issuing_authority=authority_val,
        issue_date=issue_date,
        effective_date=effective_date,
        status=status_val,
        keywords=keywords_val,
        category=category_val,
        content=content_val,
    )
    nested = db.begin_nested()
    try:
        db.add(policy)
        db.flush()
        nested.commit()
    except Exception as row_err:
        nested.rollback()
        logger.warning(f"保存政策失败 (行{row_idx}): {row_err}")
        raise row_err
    return {"ok": True}


def _parse_date_cell(row, index: int, row_idx: int, label: str) -> Optional[datetime]:
    """安全解析 Excel 单元格中的日期值。"""
    if len(row) <= index:
        return None
    try:
        val = row[index]
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            try:
                return datetime.strptime(val.strip(), "%Y-%m-%d")
            except ValueError:
                return None
    except (TypeError, AttributeError) as e:
        logger.debug(f"解析{label}失败 (行{row_idx}): {e}")
    return None


def _safe_str(row, index: int, default: Optional[str] = None) -> Optional[str]:
    """安全提取字符串字段，索引越界或类型错误时返回默认值。"""
    try:
        val = row[index] if len(row) > index and row[index] is not None else None
        return str(val).strip() if val is not None else default
    except (TypeError, ValueError, IndexError) as e:
        logger.debug(f"提取字段失败 (索引{index}): {e}")
        return default


def _make_row_error(row, row_idx: int, msg: str) -> Dict[str, Any]:
    """构造行错误字典。"""
    title = _safe_str(row, 1, default=f"未命名政策{row_idx}")
    return {"row": row_idx, "title": title, "error": msg}
