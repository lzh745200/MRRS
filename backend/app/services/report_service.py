"""
Report generation service.

Provides Excel, PDF, and comprehensive report exports for village data.
Used by data/data/reports.py route handlers.
"""

import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating data reports in various formats."""

    def __init__(self, db: Session):
        self.db = db

    async def get_reports(self, user_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """List available report templates / recent exports.

        .. warning::
            此前返回硬编码的 2 条字典（非真实数据），且无任何路由调用本方法
            （路由层直接使用 ``DataReportService`` / ``db.query`` 查询）。
            完整实现需查询真实的报表模板/导出记录表，超出本次 BugFix 范围。

        Raises:
            NotImplementedError: 本方法为硬编码占位，未接入路由。
        """
        raise NotImplementedError(
            "ReportService.get_reports 尚未实现（此前返回硬编码 2 条字典，"
            "非真实数据，未接入路由）。完整实现需查询报表模板/导出记录表，"
            "超出本次 BugFix 范围。"
        )

    async def get_report(self, report_id: int) -> Optional[Dict[str, Any]]:
        """Get a single report by ID.

        .. warning::
            此前基于 ``get_reports`` 的硬编码列表查找，非真实数据，且无路由调用。

        Raises:
            NotImplementedError: 本方法为硬编码占位，未接入路由。
        """
        raise NotImplementedError(
            "ReportService.get_report 尚未实现（此前基于硬编码列表查找，"
            "非真实数据，未接入路由）。完整实现需查询报表模板/导出记录表，"
            "超出本次 BugFix 范围。"
        )

    async def export_to_excel(self, query_params: Dict[str, Any] = None, user: Any = None) -> bytes:
        """Generate an Excel report based on query parameters.

        Args:
            query_params: Filters (year, village_ids, report_type, etc.)
            user: 当前用户，用于数据权限过滤。

        Returns:
            Excel file as bytes.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "数据报表"

            # Header
            headers = ["序号", "名称", "省份", "市县", "振兴层级", "年度", "数据值", "更新时间"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Data rows (placeholder for real data queries)
            data_rows = await self._fetch_report_data(query_params, user=user)
            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit column widths
            for col in ws.columns:
                max_length = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:  # pragma: no cover
            logger.warning("openpyxl not available, returning empty Excel")
            return self._empty_excel()
        except Exception as e:
            logger.error("Excel export failed: %s", e)
            raise

    async def export_to_pdf(self, query_params: Dict[str, Any] = None, user: Any = None) -> bytes:
        """Generate a PDF report.

        Args:
            query_params: Filters (year, village_ids, report_type, etc.)
            user: 当前用户，用于数据权限过滤。

        Returns:
            PDF file as bytes.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas

            buffer = io.BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4

            # Title
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(width / 2, height - 50, "帮扶管理信息系统 - 数据报表")

            # Metadata
            c.setFont("Helvetica", 10)
            c.drawString(50, height - 80, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            rpt_type = query_params.get("report_type", "综合报表") if query_params else "综合报表"
            c.drawString(50, height - 95, f"报表类型: {rpt_type}")

            # Data table header
            y = height - 130
            c.setFont("Helvetica-Bold", 11)
            headers = ["序号", "名称", "省份", "振兴层级"]
            x_positions = [50, 90, 250, 350]
            for header, x in zip(headers, x_positions):
                c.drawString(x, y, header)

            # Data rows
            c.setFont("Helvetica", 10)
            y -= 20
            data_rows = await self._fetch_report_data(query_params, user=user)
            for idx, row_data in enumerate(data_rows[:50], 1):  # Limit to 50 rows for PDF
                c.drawString(50, y, str(idx))
                c.drawString(90, y, str(row_data[1])[:20] if len(row_data) > 1 else "")
                c.drawString(250, y, str(row_data[2])[:15] if len(row_data) > 2 else "")
                c.drawString(350, y, str(row_data[3])[:10] if len(row_data) > 3 else "")
                y -= 18
                if y < 50:
                    c.showPage()
                    c.setFont("Helvetica", 10)
                    y = height - 50

            c.save()
            buffer.seek(0)
            return buffer.getvalue()
        except ImportError:  # pragma: no cover
            logger.warning("reportlab not available, returning empty PDF")
            return self._empty_pdf()
        except Exception as e:
            logger.error("PDF export failed: %s", e)
            raise

    async def export_comprehensive_report(
        self, year: int = None, village_ids: List[int] = None, user: Any = None
    ) -> bytes:
        """Generate a comprehensive multi-section report.

        Args:
            year: Report year.
            village_ids: List of village IDs to include.
            user: 当前用户，用于数据权限过滤。

        Returns:
            Excel file as bytes (comprehensive reports are Excel).
        """
        return await self.export_to_excel({
            "year": year or datetime.now().year,
            "village_ids": village_ids,
            "report_type": "comprehensive",
        }, user=user)

    async def get_export_filename(self, query_params: Dict[str, Any] = None) -> str:
        """Generate a suitable filename for the export.

        Args:
            query_params: Export parameters used to name the file.

        Returns:
            Filename string.
        """
        report_type = (query_params or {}).get("report_type", "report")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{report_type}_{timestamp}.xlsx"

    async def get_report_subscriptions(self, user_id: int) -> List[Dict[str, Any]]:
        """List report subscriptions for a user."""
        from app.models.supported_village import ReportSubscription

        rows = (
            self.db.query(ReportSubscription)
            .filter(ReportSubscription.user_id == user_id)
            .order_by(ReportSubscription.id.desc())
            .all()
        )
        return [
            {
                "id": r.id,
                "user_id": r.user_id,
                "name": r.name or "",
                "report_type": r.report_type or "",
                "format": r.format or "xlsx",
                "frequency": r.frequency or "weekly",
                "send_day": r.send_day,
                "send_time": r.send_time,
                "email": r.email or "",
                "output_dir": r.output_dir or "",
                "output_format": r.output_format or "pdf",
                "is_active": bool(r.is_active),
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]

    async def create_subscription(self, user_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """Create a new report subscription."""
        from app.core.transaction import safe_commit
        from app.models.supported_village import ReportSubscription

        sub = ReportSubscription(
            user_id=user_id,
            name=(data.get("name") or "").strip(),
            report_type=(data.get("report_type") or "").strip(),
            format=data.get("format") or "xlsx",
            year=data.get("year"),
            village_ids=(
                str(data["village_ids"])
                if data.get("village_ids")
                else None
            ),
            include_sections=(
                str(data["include_sections"])
                if data.get("include_sections")
                else None
            ),
            frequency=data.get("frequency") or "weekly",
            send_day=data.get("send_day"),
            send_time=data.get("send_time"),
            email=data.get("email"),
            output_dir=data.get("output_dir"),
            output_format=data.get("output_format") or "pdf",
            is_active=True,
        )
        self.db.add(sub)
        safe_commit(self.db)
        return {
            "id": sub.id,
            "user_id": sub.user_id,
            "name": sub.name,
            "report_type": sub.report_type,
            "format": sub.format,
            "frequency": sub.frequency,
            "is_active": True,
        }

    async def update_subscription(
        self, subscription_id: int, data: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Update an existing subscription."""
        from app.core.transaction import safe_commit
        from app.models.supported_village import ReportSubscription

        sub = (
            self.db.query(ReportSubscription)
            .filter(ReportSubscription.id == subscription_id)
            .first()
        )
        if sub is None:
            return None
        for key in (
            "name",
            "report_type",
            "format",
            "year",
            "village_ids",
            "include_sections",
            "frequency",
            "send_day",
            "send_time",
            "email",
            "output_dir",
            "output_format",
        ):
            if key in data and data[key] is not None:
                value = data[key]
                if key in ("village_ids", "include_sections") and isinstance(value, list):
                    value = str(value)
                setattr(sub, key, value)
        if "is_active" in data and data["is_active"] is not None:
            sub.is_active = bool(data["is_active"])
        safe_commit(self.db)
        return {
            "id": sub.id,
            "user_id": sub.user_id,
            "name": sub.name,
            "report_type": sub.report_type,
            "format": sub.format,
            "frequency": sub.frequency,
            "is_active": bool(sub.is_active),
        }

    async def cancel_subscription(self, subscription_id: int) -> bool:
        """Cancel a subscription (soft delete via is_active=False)."""
        from app.core.transaction import safe_commit
        from app.models.supported_village import ReportSubscription

        sub = (
            self.db.query(ReportSubscription)
            .filter(ReportSubscription.id == subscription_id)
            .first()
        )
        if sub is None:
            return False
        sub.is_active = False
        safe_commit(self.db)
        return True

    # ── Internal helpers ──

    async def _fetch_report_data(
        self, query_params: Dict[str, Any] = None, user: Any = None
    ) -> List[List[Any]]:
        """Fetch report data from the database.

        Args:
            query_params: 可选的查询过滤参数。
            user: 当前用户，用于数据权限过滤。

        Override this in production to query real models.
        """
        try:
            from app.core.data_permission import filter_by_data_scope
            from app.models.supported_village import SupportedVillage

            query = self.db.query(SupportedVillage)
            # 数据权限过滤（参照 villages.py:50 范式）
            query = filter_by_data_scope(query, SupportedVillage, user, db=self.db)
            rows = query.limit(100).all()
            return [
                [i + 1, r.village_name or "", r.province or "", r.county or "",
                 "是" if r.is_revitalization_tier else "否", "", "", r.updated_at.isoformat() if r.updated_at else ""]
                for i, r in enumerate(rows)
            ]
        except Exception as e:
            logger.warning("Failed to fetch report data from DB: %s", e)
            return []

    @staticmethod
    def _empty_excel() -> bytes:
        """Return a minimal empty Excel file."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.active.title = "报表"
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:  # pragma: no cover
            return b""

    @staticmethod
    def _empty_pdf() -> bytes:
        """Return a minimal empty PDF file."""
        try:
            from reportlab.pdfgen import canvas
            buffer = io.BytesIO()
            c = canvas.Canvas(buffer)
            c.drawString(100, 750, "暂无数据")
            c.save()
            buffer.seek(0)
            return buffer.getvalue()
        except ImportError:  # pragma: no cover
            return b""
