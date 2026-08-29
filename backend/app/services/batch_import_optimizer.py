"""
批量导入优化器 — pandas 加速 Excel 导入

替换 openpyxl 逐行处理为 pandas 向量化操作，
配合 SQLAlchemy bulk_insert_mappings 大幅提升导入性能。

方案 #9 — 数据导入批量化
"""

import logging
from typing import Any, Dict, List, Tuple

logger = logging.getLogger(__name__)

# 单次批量插入的批次大小
BATCH_SIZE = 500

# pandas 类型映射（dtype=str 可避免科学记数法等转换问题）
PANDAS_READ_KWARGS = {
    "engine": "openpyxl",
    "dtype": str,
    "keep_default_na": False,
}


def read_excel_raw(file_content: bytes) -> List[List[Any]]:
    """pandas 快速读取全部原始行（header=None，不做任何表头假设）。

    供 ExcelImporterService.parse_excel 使用：官方模板的表头不在第 1 行
    （前 5 行为装饰标题区），表头探测与列名映射由调用方完成。

    Returns:
        行列表，每行为单元格值列表（空值为 None，与 openpyxl values_only 一致）
    """
    import pandas as pd
    from io import BytesIO

    df = pd.read_excel(
        BytesIO(file_content),
        engine="openpyxl",
        header=None,
        dtype=object,
        keep_default_na=False,
    )
    # keep_default_na=False 会把空单元格读成 ""，统一回 None 与 openpyxl 对齐
    return [[(None if v == "" else v) for v in row] for row in df.values.tolist()]


def _read_excel_fallback(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """openpyxl 回退读取（pandas 不可用时）"""
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 第一行是表头
    try:
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        wb.close()
        return [], []

    rows = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)})

    wb.close()
    return rows, headers
